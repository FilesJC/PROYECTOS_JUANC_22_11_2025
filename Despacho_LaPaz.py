import plotly.express as px
import pandas as pd
import os
import warnings
import altair as alt
import plotly.graph_objects as go
import xlsxwriter
from pathlib import Path
import glob
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime


warnings.filterwarnings('ignore')

# https://www.webfx.com/tools/emoji-cheat-sheet/

st.set_page_config(page_title="Sistema Almacen - Despacho La Paz!!!", page_icon=":package:",layout="wide")

 #------********************* RUTA DE IMAGEN **********************************-----------
 # Logo 
st.sidebar.image("C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Dashboard_Poly/Poly_logo.png", use_column_width=True, caption="SISTEMA ALMACEN LA PAZ")
#------***********************************************************************------------

st.sidebar.header("SELECCIONES UNA OPCION: ")
proyectos = st.sidebar.selectbox(
        "Opciones",
        options=["OPCIONES", "FILTRADOR ALMACEN", "CREAR GUIA RUTA LA PAZ", "JUEGA CON DINO"]
    )

if proyectos == "OPCIONES":
   
    st.title(" :card_index_dividers: SISTEMA ALMACEN LA PAZ - BOLIVIA")
    st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

    # Mostrar el total de ingresos centrado en una caja
    st.markdown(
        f"""
        <div style='
            border: 5px solid black; 
            padding: 5px; 
            border-radius: 2px; 
            text-align: center; 
            background-color: #6e0000;'>
            <h3 style='color: white;'>BIENVENIDO AL SISTEMA ALMACEN LA PAZ BOLIVIA</h3>
            
        </div>
        """, 
        unsafe_allow_html=True
    )

    #------********************* RUTA DE IMAGEN **********************************-----------
    # Mostrar la imagen de portada
    st.image("C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Dashboard_Poly/Presentacion.png", use_column_width=True, 
    caption="")

    st.markdown("""
    <div style="width: 100%; overflow: hidden; white-space: nowrap;">
    <div style="
        display: inline-block;
        padding-left: 100%;
        animation: scroll-left 10s linear infinite;
        font-size: 18px;
        color: #3498db;
    ">
        DERECHOS-RECERVADOS - @JUAN CARLOS RAMOS CHURA - 2025
    </div>
    </div>

    <style>
    @keyframes scroll-left {
    0%   { transform: translateX(0%); }
    100% { transform: translateX(-100%); }
    }
    </style>
    """, unsafe_allow_html=True)

    #------*************************************************************************-----------

if proyectos == "FILTRADOR ALMACEN":
    import streamlit as st
    import pandas as pd
    from openpyxl import Workbook
    from io import BytesIO
    import xlsxwriter
    from pathlib import Path
    import os
    import glob

    #st.set_page_config(layout="wide")

    st.markdown("<h1 style='text-align: center;'>FILTRADO DE DOCUMENTOS Y CAJAS ALMACEN</h1>", unsafe_allow_html=True)

    opciones = ["SELECCIONA UNA OPCION","FILTRADO FILES", "FILTRADO TOMOS", "FILTRADO DE CAJAS"]

    seleccion = st.selectbox("📋 Selecciona una opcion del menu: ", opciones)


    if seleccion == "SELECCIONA UNA OPCION":   
        mensaje_markdown = """
        ### BIENVENIDO !! FILTRADO DE DOCUMENTOS ALMACEN

        **Desarrollado por Juan Carlos Ramos Chura**
        """
        st.markdown(mensaje_markdown)

    elif seleccion == "FILTRADO FILES":


        col1, col2 = st.columns(2)

        with col1:
            st.markdown("<h2 style='text-align: center;'> 📤 CARGAR PLANILLA DE EXCEL PARA FILTRAR FILES</h2>", unsafe_allow_html=True)

            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])

            if uploaded_file is not None:
            
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)
                
                #ruta = st.text_input("Introduce la ruta de la carpeta: Por Ejemplo", "C:/Users/juan.ramos/Desktop/")
        
                st.write('🔍 FILTRADO POR NIVELES:')

                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES","Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])
                
                Nivel = st.selectbox("Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":

                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")
                        
                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                if Nivel == "2":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                if Nivel == "3":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


                if Nivel == "4":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")
                        
                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                        
                if Nivel == "5":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        
                if Nivel == "6":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        
                # Mostrar un mensaje
                st.write('🔍 FILTRADO POR LOCACIONES:')

                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":
                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    st.dataframe(Ordenar)

                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones
            #st.write("Sube varios archivos Excel para combinarlos en uno solo.")
            st.markdown("<h2 style='text-align: center;'>📎 SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar múltiples archivos
            uploaded_files = st.file_uploader("🔍 Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Función para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="📥 Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Files.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")


            st.markdown("<h2 style='text-align: center;'>🗑️ ELIMINAR ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Subir archivos Excel
            uploaded_files = st.file_uploader(
                "Selecciona uno o varios archivos Excel para procesar o eliminar:",
                type=["xls", "xlsx"],
                accept_multiple_files=True
            )

            if uploaded_files:
                st.write(f"Se subieron {len(uploaded_files)} archivos.")

                if st.button("❌ Eliminar archivos subidos"):
                    uploaded_files.clear()  # Esto elimina los archivos del buffer de Streamlit
                    st.success("✅ Los archivos han sido eliminados de la sesión actual.")

            else:
                st.info("Haz clic para seleccionar archivos. No se mostrará ninguna ruta.")            
            

            # ---------------------------------------------------------------------------------------------------------

    elif seleccion == "FILTRADO TOMOS":

        col1, col2 = st.columns(2)
        with col1:
            # Titulo de Aplicacion
        
            st.markdown("<h2 style='text-align: center;'>📤 CARGAR PLANILLA DE EXCEL PARA FILTRAR TOMOS</h2>", unsafe_allow_html=True)

        
            #Cargar el archivo de excel 
            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])

        
            if uploaded_file is not None:
                # Leer el archivo Excel usando Pandas
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                # Elimoinar Columnas
                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)

                # Definimos una ruta para guardar nuestros archivos
                #ruta = st.text_input("Introduce la ruta de la carpeta: Por Ejemplo", "C:/Users/juan.ramos/Desktop/")
                
            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("📁 Guardar Como:", options = ["OPCIONES", "Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])

                Nivel = st.selectbox("🔍 Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        
                if Nivel == "2":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                        
                if Nivel == "3":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    

                if Nivel == "4":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                        
                if Nivel == "5":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                        
                if Nivel == "6":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":
                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    st.dataframe(Ordenar)

                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones
            st.markdown("<h2 style='text-align: center;'>📎 SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar múltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Función para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="📥 Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Tomos.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")

            # Título de la aplicación

            st.markdown("<h2 style='text-align: center;'>🗑️ ELIMINAR ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Subir archivos Excel
            uploaded_files = st.file_uploader(
                "Selecciona uno o varios archivos Excel para procesar o eliminar:",
                type=["xls", "xlsx"],
                accept_multiple_files=True
            )

            if uploaded_files:
                st.write(f"Se subieron {len(uploaded_files)} archivos.")

                if st.button("❌ Eliminar archivos subidos"):
                    uploaded_files.clear()  # Esto elimina los archivos del buffer de Streamlit
                    st.success("✅ Los archivos han sido eliminados de la sesión actual.")

            else:
                st.info("Haz clic para seleccionar archivos. No se mostrará ninguna ruta.")            
            
        
            # ---------------------------------------------------------------------------------------------------------


    elif seleccion == "FILTRADO DE CAJAS":

        col1, col2 = st.columns(2)
        with col1:
            # Titulo de Aplicacion
        
            st.markdown("<h2 style='text-align: center;'>📤 CARGAR PLANILLA DE EXCEL PARA FILTRAR CAJAS</h2>", unsafe_allow_html=True)
        
            #Cargar el archivo de excel 
            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])

        
            if uploaded_file is not None:
                # Leer el archivo Excel usando Pandas
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                # Elimoinar Columnas
                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)

                # Definimos una ruta para guardar nuestros archivos
                #ruta = st.text_input("Introduce la ruta de la carpeta: Por Ejemplo", "C:/Users/juan.ramos/Desktop/")
            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("📂 Guardar Como:", options = ["OPCIONES", "Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])

                Nivel = st.selectbox("🔍 Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")
                        
                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                if Nivel == "2":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        
                if Nivel == "3":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        
                if Nivel == "4":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


                if Nivel == "5":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


                if Nivel == "6":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001", "L-PREDESP_IN", "L-PREDESP_EX", "L-ING-CJ-001", "L-INV-CJ-001", "L-SCN-CJ-001", "L-DIG-CJ-001"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV", "PREDESP_IN", "PREDESP_EX", "ING", "INV", "SCN", "DIG"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        
                if Loc == "L-PREDESP_IN":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PREDESP_IN" and Loc == "PREDESP_IN":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")
                        
                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                if Loc == "L-PREDESP_EX":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PREDESP_EX" and Loc == "PREDESP_EX":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                if Loc == "L-ING-CJ-001":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-ING-CJ-001" and Loc == "ING":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        
                if Loc == "L-INV-CJ-001":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-INV-CJ-001" and Loc == "INV":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                if Loc == "L-SCN-CJ-001":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-SCN-CJ-001" and Loc == "SCN":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                if Loc == "L-DIG-CJ-001":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-DIG-CJ-001" and Loc == "DIG":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Guardar en memoria
                        output = BytesIO()

                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 
                        
                        output.seek(0)

                        # Botón de descarga directa
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=output,
                            file_name=f"{file_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones       
            st.markdown("<h2 style='text-align: center;'>📎 SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar múltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Función para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="📥 Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Cajas.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")


            
            st.markdown("<h2 style='text-align: center;'>🗑️ ELIMINAR ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Subir archivos Excel
            uploaded_files = st.file_uploader(
                "Selecciona uno o varios archivos Excel para procesar o eliminar:",
                type=["xls", "xlsx"],
                accept_multiple_files=True
            )

            if uploaded_files:
                st.write(f"Se subieron {len(uploaded_files)} archivos.")

                if st.button("❌ Eliminar archivos subidos"):
                    uploaded_files.clear()  # Esto elimina los archivos del buffer de Streamlit
                    st.success("✅ Los archivos han sido eliminados de la sesión actual.")

            else:
                st.info("Haz clic para seleccionar archivos. No se mostrará ninguna ruta.")            
            

if proyectos == "CREAR GUIA PARA RUTA":

    import streamlit as st
    import pandas as pd
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from datetime import datetime

    st.title("📦 CREAR GUIA DE RUTA PARA MENSAJERIA - POLYSISTEMAS - LA PAZ")

    # Cargar archivos
    col1, col2, col3 = st.columns(3)
    with col1:
        uploaded_file = st.file_uploader("Cargar archivo combinado con Items Oneil", type="xlsx")
    with col2:
        logo_file = st.file_uploader("📷 Cargar primer logo", type=["png", "jpg", "jpeg"])
    with col3:
        logo_file2 = st.file_uploader("📷 Cargar segundo logo", type=["png", "jpg", "jpeg"], key="logo2")

    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip().str.replace(" ", "_").str.replace("-", "_")

        required_columns = ['Solicitante', 'Centro_de_Costo', 'WorkOrderCode', 'Cantidad', 'Items_Oneil', 'TipoFile']
        if not all(col in df.columns for col in required_columns):
            st.error(f"❌ Faltan columnas requeridas: {', '.join(required_columns)}")
            st.stop()

        colum1, colum2, colum3 = st.columns(3)
        with colum1:
            # Configuración de filtros y datos
            tipos_disponibles = df['TipoFile'].dropna().unique()
            tipos_seleccionados = st.multiselect("Selecciona uno o más TipoFile a procesar:", tipos_disponibles)
            if not tipos_seleccionados:
                st.warning("⚠️ Por favor, selecciona al menos un TipoFile para continuar.")
                st.stop()
        with colum2:
            df = df[df['TipoFile'].isin(tipos_seleccionados)]

            encargado = st.selectbox("👤 Encargado:", ["", "JAIME QUISPE", "CARLOS ORTIZ", "MARCO HUAYLLUCO", "ALFREDO RIVEROS"])
            if not encargado:
                st.warning("⚠️ Por favor, selecciona el encargado.")
                st.stop()
        with colum3:
            guia_seleccionada = st.selectbox("Tipo de Guía:", ["GUÍA DE RECEPCIÓN DE CAJAS", "GUÍA DE RECOJO DE ITEMS", "GUÍA DE ENTREGA DE MATERIALES"])
            regional_seleccionada = st.selectbox("Regional:", ["REGIONAL LA PAZ", "REGIONAL EL ALTO", "REGIONAL SUCRE", "REGIONAL ORURO"])
            fecha_actual = datetime.now().strftime("%d/%m/%Y")

        st.info(f"""
            **👤 Encargado:** {encargado}  
            **📅 Fecha del Reporte:** {fecha_actual}  
            **📋 Tipo de Guía:** {guia_seleccionada}  
            **📍 Regional:** {regional_seleccionada}
            """)

        st.subheader("🔍 Vista previa de datos")
        st.write(df.head(2))

        # Preparar workbook
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Detallado"
        ws.sheet_view.showGridLines = False

        # Estilos
        bold = Font(bold=True)
        verde = PatternFill("solid", fgColor="A9D18E")
        morado = PatternFill("solid", fgColor="C27BA0")
        gris = PatternFill("solid", fgColor="D9D9D9")
        center = Alignment(horizontal="center", vertical="center")
        wrap_text = Alignment(wrap_text=True, vertical="top")
        thin = Side(style="thin")
        thick = Side(style="thick")

        # Logos
        if logo_file:
            image_stream = BytesIO(logo_file.read())
            logo = OpenpyxlImage(image_stream)
            logo.width, logo.height = 190, 80
            ws.add_image(logo, "A1")
        if logo_file2:
            image_stream2 = BytesIO(logo_file2.read())
            logo2 = OpenpyxlImage(image_stream2)
            logo2.width, logo2.height = 190, 80
            ws.add_image(logo2, "G1")

        # ENCABEZADO ENMARCADO
        encabezados = [
            (1, guia_seleccionada),
            (2, regional_seleccionada),
            (3, "RESPONSABLE POLYSISTEMAS BOLIVIA S.R.L.")
        ]
        for row, texto in encabezados:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.cell(row=row, column=1, value=texto).font = bold
            ws.cell(row=row, column=1).alignment = center
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = Border(
                    top=thick if row == 1 else thin,
                    bottom=thick if row == 3 else thin,
                    left=thick if col == 1 else thin,
                    right=thick if col == 7 else thin
                )

        # Fila con ENCARGADO y FECHA
        ws.merge_cells("A5:B5")
        ws.merge_cells("C5:D5")
        ws.merge_cells("E5:F5")
        ws["A5"].value = "ENCARGADO:"
        ws["C5"].value = encargado
        ws["E5"].value = "FECHA:"
        ws["G5"].value = fecha_actual
        for col in [1, 3, 5, 7]:
            ws.cell(row=5, column=col).font = bold
            ws.cell(row=5, column=col).alignment = center
        for col in range(1, 8):
            ws.cell(row=5, column=col).border = Border(
                top=thin,
                bottom=thick,
                left=thick if col == 1 else thin,
                right=thick if col == 7 else thin
            )

        # Total cajas
        ws.merge_cells("A6:D6")
        ws["A6"] = "CANTIDAD TOTAL"
        ws["A6"].font = bold
        ws["A6"].alignment = center
        ws["E6"] = df['Cantidad'].sum()
        ws["E6"].font = bold
        ws["E6"].alignment = center
        for col in range(1, 6):
            ws.cell(row=6, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

        current_row = 7
        df = df.dropna(subset=['Solicitante'])

        # Agrupar por Solicitante
        for solicitante, grupo_solicitante in df.groupby('Solicitante'):
            for centro, grupo_agencia in grupo_solicitante.groupby('Centro_de_Costo'):
                start_row = current_row
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
                ws.cell(row=current_row, column=1, value=f"SOLICITANTE: {solicitante}").font = bold
                ws.cell(row=current_row, column=5, value=f"AGENCIA: {centro}").font = bold
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).fill = verde
                    ws.cell(row=current_row, column=col).alignment = center
                    ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                current_row += 1

                headers = ["NRO", "NRO. WO", "CANTIDAD", "CONTAINER CODE", "OBSERVACIONES"]
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=current_row, column=col_num, value=header).font = bold
                    ws.cell(row=current_row, column=col_num).fill = morado
                    ws.cell(row=current_row, column=col_num).alignment = center
                    ws.cell(row=current_row, column=col_num).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                current_row += 1

                total_cantidad = 0
                for i, row in enumerate(grupo_agencia.itertuples(index=False), 1):
                    ws.cell(row=current_row, column=1, value=i).alignment = center
                    ws.cell(row=current_row, column=2, value=row.WorkOrderCode).alignment = center
                    ws.cell(row=current_row, column=3, value=row.Cantidad).alignment = center
                    ws.cell(row=current_row, column=4, value=str(row.Items_Oneil).replace(";", ",")).alignment = wrap_text
                    ws.cell(row=current_row, column=5, value="").alignment = wrap_text
                    for col in range(1, 6):
                        ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    total_cantidad += row.Cantidad
                    current_row += 1

                # Totales y Firmas
                ws.cell(row=current_row, column=2, value="TOTAL").font = bold
                ws.cell(row=current_row, column=3, value=total_cantidad).font = bold
                for col in range(1, 6):
                    ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

                ws.cell(row=current_row, column=6, value="ENTREGUE CONFORME").font = bold
                ws.cell(row=current_row, column=6).fill = gris
                ws.cell(row=current_row, column=6).alignment = center
                ws.cell(row=current_row, column=6).border = Border(left=thin, right=thin, top=thin, bottom=thin)

                ws.cell(row=current_row, column=7, value="RECIBÍ CONFORME").font = bold
                ws.cell(row=current_row, column=7).fill = gris
                ws.cell(row=current_row, column=7).alignment = center
                ws.cell(row=current_row, column=7).border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Enmarcar todo el bloque con borde grueso
                end_row = current_row
                for r in range(start_row, end_row + 1):
                    for c in range(1, 8):
                        cell = ws.cell(row=r, column=c)
                        cell.border = Border(
                            top=thick if r == start_row else thin,
                            bottom=thick if r == end_row else thin,
                            left=thick if c == 1 else thin,
                            right=thick if c == 7 else thin
                        )
                current_row += 2

        # Ajuste de columnas
        for i, width in enumerate([8, 15, 10, 50, 30, 25, 25], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(output)
        output.seek(0)

        nombre_archivo = st.text_input("📝 Ingresa el nombre del archivo:", "Reporte_Solicitantes")
        if not nombre_archivo.strip():
            st.warning("⚠️ Por favor, ingresa un nombre válido para el archivo.")
        else:
            st.success("✅ Reporte generado con encabezado completo, firmas y totales.")
            st.download_button(
            label="📥 Descargar Reporte Final",
            data=output,
            file_name=f"{nombre_archivo.strip()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


if proyectos == "CREAR GUIA RUTA LA PAZ":

    import streamlit as st
    import pandas as pd
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from datetime import datetime

    st.title("📦 GENERADOR DE GUÍA - FORMATO OFICIAL POLYSISTEMAS")

    # Cargar archivos
    col1, col2 = st.columns(2)
    with col1:
        # Subir imagen del logo
        logo_file = st.file_uploader("📸 Subir Logo (PNG o JPG)", type=["png", "jpg", "jpeg"])
    with col2:
        # Subir archivo de datos
        uploaded_file = st.file_uploader("Cargar archivo Excel para generar la Ruta", type="xlsx")

    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip().str.replace(" ", "_").str.replace("-", "_")

        required_columns = ['Solicitante', 'Centro_de_Costo', 'WorkOrderCode', 'Cantidad', 'Items_Oneil', 'TipoFile']
        if not all(col in df.columns for col in required_columns):
            st.error(f"❌ Faltan columnas requeridas: {', '.join(required_columns)}")
            st.stop()

        # Selectores
        col1, col2, col3 = st.columns(3)

        with col1:
            personal = st.selectbox("👤 Personal de Polysistemas:",
                                    ["JAIME QUISPE", "CARLOS ORTIZ", "MARCO HUAYLLUCO", "ALFREDO RIVEROS"])

        with col2:
            cliente = st.selectbox("🏢 Cliente:",
                                ["BANCO SOL", "BNB", "BANCO FIE"])

        with col3:
            fecha_actual = datetime.now().strftime("%d/%m/%Y")

            tipo_seleccionado = st.multiselect(
                "Selecciona uno o más TipoFile a procesar:",
                ["CAJA", "Caja", "Cintillos", "FILE"],
                default=["CAJA"]
            )

        # Filtrar TipoFile
        df = df[df["TipoFile"].isin(tipo_seleccionado)]

        st.info(f"""
        **👤 Personal:** {personal}  
        **🏢 Cliente:** {cliente}  
        **📅 Fecha:** {fecha_actual}  
        **🎯 TipoFile Seleccionado:** {', '.join(tipo_seleccionado)}  
        """)
        
        st.subheader("🔍 Vista previa de datos filtrados")
        st.write(df.head(5))

        # Crear Excel
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Guía"
        ws.sheet_view.showGridLines = False

        # Estilos
        bold = Font(bold=True)
        white_bold = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        wrap = Alignment(wrap_text=True, vertical="top")

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        rojo = PatternFill("solid", fgColor="933C47")
        gris = PatternFill("solid", fgColor="F2F2F2")

        # FUNCIÓN: aplicar bordes a toda una fila
        def border_full_row(row, cols=7):
            for c in range(1, cols + 1):
                ws.cell(row, c).border = border

        # Insertar logo
        if logo_file:
            img = XLImage(logo_file)
            img.width = 190
            img.height = 75
            ws.add_image(img, "A1")
            
        # Estilos de borde (asegúrate que esté definido antes o ponlo aquí)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # --------------------------
        # ENCABEZADO
        # --------------------------
        ws.merge_cells("C1:E1")
        ws["C1"].value = "FORMATO"
        ws["C1"].font = bold
        ws["C1"].alignment = center
        for col in range(3, 6):  # columnas C(3) a E(5)
            ws.cell(row=1, column=col).border = border
        

        ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=5)
        ws["C2"].value = "GUIA DE RECEPCIÓN DE MENSAJERÍA"
        ws["C2"].font = bold
        ws["C2"].alignment = center
        for col in range(3, 6):
            ws.cell(row=2, column=col).border = border
        for col in range(3, 6):
            ws.cell(row=3, column=col).border = border

        encabezado_pairs = {
            "F1": "Código:", "G1": "BOL-2.18-FOR-01",
            "F2": "Edición:", "G2": "01",
            "F3": "Fecha Edición:", "G3": "4/11/2025"
        }

        for cell, value in encabezado_pairs.items():
            ws[cell].value = value
            ws[cell].alignment = center
            ws[cell].border = border
            if "F" in cell:
                ws[cell].font = bold

        # --------------------------
        # PERSONAL / FECHA
        # --------------------------
        ws.merge_cells("A5:C5")
        ws["A5"].value = "PERSONAL DE POLYSISTEMAS"
        ws["A5"].font = white_bold
        ws["A5"].fill = rojo
        ws["A5"].alignment = left
        border_full_row(5)

        ws.merge_cells("D5:E5")
        ws["D5"].value = personal
        ws["D5"].alignment = center
        ws["D5"].font = bold

        ws["F5"].value = "FECHA:"
        ws["F5"].font = white_bold
        ws["F5"].fill = rojo
        ws["F5"].alignment = left

        ws["G5"].value = fecha_actual
        ws["G5"].alignment = center
        ws["G5"].font = bold

        border_full_row(5)

        # CLIENTE
        ws.merge_cells("A6:C6")
        ws["A6"].value = "CLIENTE:"
        ws["A6"].font = white_bold
        ws["A6"].fill = rojo
        ws["A6"].alignment = left

        ws.merge_cells("D6:E6")
        ws["D6"].value = cliente
        ws["D6"].alignment = center
        ws["D6"].border = border
        ws["D6"].font = bold

        ws["F6"].value = "CANTIDAD TOTAL:"
        ws["F6"].font = white_bold
        ws["F6"].fill = rojo
        ws["F6"].alignment = left

        total_cantidad = df["Cantidad"].sum()
        ws["G6"].value = total_cantidad
        ws["G6"].alignment = center
        ws["G6"].font = bold

        border_full_row(6)

        # --------------------------
        # TABLA POR SOLICITANTE
        # --------------------------
        solicitantes_unicos = df["Solicitante"].unique()
        start_row = 8

        for solicitante in solicitantes_unicos:

            df_s = df[df["Solicitante"] == solicitante]
            agencia = df_s["Centro_de_Costo"].iloc[0]

            # FILA DE SOLICITANTE
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
            ws.cell(start_row, 1).value = "SOLICITANTE:"
            ws.cell(start_row, 1).fill = gris
            ws.cell(start_row, 1).font = bold
            ws.cell(start_row, 1).alignment = left

            ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=5)
            ws.cell(start_row, 3).value = solicitante
            ws.cell(start_row, 3).font = bold
            ws.cell(start_row, 3).alignment = center

            ws.cell(start_row, 6).value = "AGENCIA:"
            ws.cell(start_row, 6).fill = gris
            ws.cell(start_row, 6).font = bold
            ws.cell(start_row, 6).alignment = left

            ws.cell(start_row, 7).value = agencia
            ws.cell(start_row, 7).font = bold
            ws.cell(start_row, 7).alignment = center

            border_full_row(start_row)
            start_row += 1

            headers = ["NRO", "NRO. WO", "CANTIDAD", "CONTAINER CODE", "OBSERVACIONES", "FIRMA ENTREGA", "FIRMA RECEPCIÓN"]
            for col, h in enumerate(headers, 1):
                ws.cell(start_row, col).value = h
                ws.cell(start_row, col).font = white_bold
                ws.cell(start_row, col).fill = rojo
                ws.cell(start_row, col).alignment = center
            border_full_row(start_row)
            start_row += 1

            nro = 1
            total_solic = 0
            for r in df_s.itertuples(index=False):
                ws.cell(start_row, 1, nro).alignment = center
                ws.cell(start_row, 2, r.WorkOrderCode).alignment = center
                ws.cell(start_row, 3, r.Cantidad).alignment = center
                ws.cell(start_row, 4, str(r.Items_Oneil).replace(";", ",")).alignment = center
                border_full_row(start_row)

                total_solic += r.Cantidad
                nro += 1
                start_row += 1

            # TOTAL + FIRMAS
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
            ws.cell(start_row, 1).value = "TOTAL"
            ws.cell(start_row, 1).font = bold
            ws.cell(start_row, 1).alignment = center

            ws.cell(start_row, 3).value = total_solic
            ws.cell(start_row, 3).alignment = center
            
            ws.cell(start_row, 6).value = "ENTREGUÉ CONFORME"
            ws.cell(start_row, 6).fill = gris
            ws.cell(start_row, 6).font = bold
            ws.cell(start_row, 6).alignment = center

            ws.cell(start_row, 7).value = "RECIBÍ CONFORME"
            ws.cell(start_row, 7).fill = gris
            ws.cell(start_row, 7).font = bold
            ws.cell(start_row, 7).alignment = center

            border_full_row(start_row)
            start_row += 2  # ← YA NO DEJA SALTOS VACÍOS

        # Ajustar ancho columnas
        widths = [10, 20, 12, 40, 30, 22, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(output)
        output.seek(0)

        st.download_button(
            label="📥 Descargar Guía en Formato Oficial",
            data=output,
            file_name="Guia_Polysistemas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


if proyectos == "JUEGA CON DINO":

    import streamlit as st

    st.title("🦖 Dinosaurio Saltarín ☁️")

    # Botón de reinicio
    if st.button("🔄 Reiniciar Juego"):
        st.rerun()

    game_code = """
    <style>
    html, body {
        margin: 0;
        padding: 0;
        height: 100%;
        background-color: black;
    }

    #gameContainer {
        width: 100%;
        height: 100%;
        background-color: black;
    }

    canvas {
        display: block;
        width: 100%;
        height: 100%;
        background-color: #cce7ff;  /* Día por defecto */
        border: 2px solid #333;
    }
    </style>

    <div id="gameContainer">
    <canvas id="gameCanvas"></canvas>
    </div>

    <script>
    const canvas = document.getElementById("gameCanvas");
    const ctx = canvas.getContext("2d");

    function resizeCanvas() {
        canvas.width = canvas.clientWidth;
        canvas.height = canvas.clientHeight;
    }
    resizeCanvas();
    window.addEventListener("resize", resizeCanvas);

    ctx.textBaseline = "bottom";

    let groundLevel = canvas.height - 20;

    let dino = {
        x: 50, y: groundLevel - 40, width: 40, height: 40,
        dy: 0, gravity: 1, jumpPower: -15, grounded: true
    };

    let obstacles = [
        { x: 800, width: 20, height: 40, type: "smallCactus", speed: 6 },
        { x: 1200, width: 30, height: 60, type: "bigCactus", speed: 6 },
        { x: 1600, width: 30, height: 30, type: "bird", speed: 8 }
    ];

    // Ajustar posición inicial
    obstacles.forEach(o => {
        if (o.type === "bird") {
            o.y = groundLevel - o.height - 60;
        } else {
            o.y = groundLevel - o.height;
        }
    });

    // Nubes decorativas ☁️
    let clouds = [
        { x: 100, y: 50, speed: 1 },
        { x: 400, y: 80, speed: 0.8 },
        { x: 700, y: 60, speed: 1.2 }
    ];

    function drawClouds() {
        ctx.font = "30px Arial";
        clouds.forEach(c => {
            ctx.fillText("☁️", c.x, c.y);
        });
    }

    function updateClouds() {
        clouds.forEach(c => {
            c.x -= c.speed;
            if (c.x < -50) {
                c.x = canvas.width + Math.random() * 200;
                c.y = 30 + Math.random() * 60;
            }
        });
    }

    // Día y noche 🌞🌙
    let isDay = true;

    function updateBackgroundColor() {
        if (score % 20 === 0 && score !== 0 && score !== lastToggleScore) {
            isDay = !isDay;
            lastToggleScore = score;
        }

        if (isDay) {
            canvas.style.backgroundColor = "#cce7ff";  // Día
        } else {
            canvas.style.backgroundColor = "#2c3e50";  // Noche
        }
    }

    let gameOver = false;
    let score = 0;
    let lastToggleScore = -1;

    function drawBackground() {
        ctx.fillStyle = "#8B4513";
        ctx.fillRect(0, canvas.height - 20, canvas.width, 20);
    }

    function drawDino() {
        ctx.font = "40px Arial";
        ctx.fillText("🦖", dino.x, dino.y + dino.height);
    }

    function drawObstacles() {
        ctx.font = "40px Arial";
        obstacles.forEach(o => {
            if (o.type === "smallCactus" || o.type === "bigCactus") {
                ctx.fillText("🌵", o.x, o.y + o.height);
            } else if (o.type === "bird") {
                ctx.fillText("🐦", o.x, o.y + o.height);
            }
        });
    }

    function update() {
        if (gameOver) return;

        ctx.clearRect(0, 0, canvas.width, canvas.height);

        updateBackgroundColor();  // Cambiar día/noche
        updateClouds();
        drawBackground();
        drawClouds();

        groundLevel = canvas.height - 20;

        // Dino saltando
        dino.y += dino.dy;
        if (dino.y + dino.height < groundLevel) {
            dino.dy += dino.gravity;
            dino.grounded = false;
        } else {
            dino.y = groundLevel - dino.height;
            dino.dy = 0;
            dino.grounded = true;
        }

        // Obstáculos
        obstacles.forEach(o => {
            o.x -= o.speed;
            if (o.x + o.width < 0) {
                o.x = canvas.width + Math.random() * 400;

                if (o.type === "bird") {
                    o.y = groundLevel - o.height - 60;
                } else {
                    o.y = groundLevel - o.height;
                }

                score++;
                if (score % 5 === 0) o.speed += 1;
            }

            // Colisión
            if (
                dino.x < o.x + o.width &&
                dino.x + dino.width > o.x &&
                dino.y < o.y + o.height &&
                dino.y + dino.height > o.y
            ) {
                gameOver = true;
                alert("💀 Game Over! Fuck you Puntuación final: " + score);
            }
        });

        drawDino();
        drawObstacles();

        ctx.fillStyle = isDay ? "black" : "white";
        ctx.font = "20px Arial";
        ctx.fillText("Score: " + score, 10, 30);

        requestAnimationFrame(update);
    }

    window.addEventListener("keydown", function(e) {
        if ((e.code === "Space" || e.code === "ArrowUp") && dino.grounded) {
            dino.dy = dino.jumpPower;
        }
    });

    update();
    </script>
    """

    st.components.v1.html(game_code, height=450)


