# Intalar tkinter mas por favor
#pip install tk
import streamlit as st
import plotly.express as px
import pandas as pd
import os
import warnings
import altair as alt
import plotly.graph_objects as go
from openpyxl import Workbook
from io import BytesIO
import xlsxwriter
from pathlib import Path
import glob


warnings.filterwarnings('ignore')

# https://www.webfx.com/tools/emoji-cheat-sheet/

st.set_page_config(page_title="Sistema Almacen La Paz!!!", page_icon=":card_file_box:",layout="wide")

 #------********************* RUTA DE IMAGEN **********************************-----------
 # Logo 
st.sidebar.image("C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Dashboard_Poly/Poly_logo.png", use_column_width=True, caption="SISTEMA ALMACEN LA PAZ")
#------***********************************************************************------------

st.sidebar.header("SELECCIONES UNA OPCION: ")
proyectos = st.sidebar.selectbox(
        "Opciones",
        options=["OPCIONES", "CONTROL ALMACEN", "FILTRADOR ALMACEN", "CALENDARIO ALMACEN", "DASHBOARD ALMACEN", "DASHBOARD DESPACHO", "CREAR RUTA", "CHECK FILEWEB AND LASERFICHE", "SINTAXIS LASERFICHE-ONEIL", "SCRIPTS"]
    )


if proyectos == "OPCIONES":
   
    st.title(" :card_index_dividers: SISTEMA ALMACEN LA PAZ - BOLIVIA")
    st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

    # Mostrar el total de ingresos centrado en una caja
    st.markdown(
        f"""
        <div style='
            border: 5px solid black; 
            padding: 5px; 
            border-radius: 2px; 
            text-align: center; 
            background-color: #6e0000;'>
            <h3 style='color: white;'>BIENVENIDO AL SISTEMA ALMACEN LA PAZ BOLIVIA</h3>
            
        </div>
        """, 
        unsafe_allow_html=True
    )

    #------********************* RUTA DE IMAGEN **********************************-----------
    # Mostrar la imagen de portada
    st.image("C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Dashboard_Poly/Presentacion.png", use_column_width=True, 
    caption="")

    st.markdown("""
    <div style="width: 100%; overflow: hidden; white-space: nowrap;">
    <div style="
        display: inline-block;
        padding-left: 100%;
        animation: scroll-left 10s linear infinite;
        font-size: 18px;
        color: #3498db;
    ">
        DERECHOS-RECERVADOS - @JUAN CARLOS RAMOS CHURA - 2025
    </div>
    </div>

    <style>
    @keyframes scroll-left {
    0%   { transform: translateX(0%); }
    100% { transform: translateX(-100%); }
    }
    </style>
    """, unsafe_allow_html=True)

    #------*************************************************************************-----------

elif proyectos == "CONTROL ALMACEN":
    
    st.title(" :desktop_computer: CONTROL ALMACEN")
    st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

    # Crear columnas con tamaños personalizados (proporcionales)
    col1, col2, col3, col4 = st.columns(4)  # La segunda columna será más ancha

    # Colocar las imágenes en las columnas
    with col1:  
        #------********************* RUTA DE IMAGEN **********************************-----------
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/Nuevo_Ingreso.png"
        #------************************************************************************-----------
        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=250)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #9932CC; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                    }
                        
                .stButton>button:hover {
                    background-color: #9932CC; /* Color de fondo al pasar el mouse */
                    }
            </style>
        """, unsafe_allow_html=True)
        #--------************** CAMBIAR RUTA ********************---------
        # Especifica la ruta de la carpeta que deseas listar
        carpeta = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/GUIAS_RECEPCION"  # Cambia esto a la ruta de tu carpeta
        #--------************** CAMBIAR RUTA ********************---------

        # Inicializa una variable de sesión para almacenar los archivos
        if 'archivos' not in st.session_state:
            st.session_state.archivos = []

        # Función para listar archivos
        def listar_archivos():
            # Listar archivos en la carpeta
            st.session_state.archivos = os.listdir(carpeta)
            # Filtrar solo archivos (opcional)
            st.session_state.archivos = [archivo for archivo in st.session_state.archivos if os.path.isfile(os.path.join(carpeta, archivo))]

        # Aplicar estilos CSS personalizados al botón


        # Botón para cargar archivos
        if st.button("INGRESO GENERAL"):
            listar_archivos()
            #st.success("Archivos cargados.")

        # Mostrar los archivos en un selectbox si hay archivos disponibles
        if st.session_state.archivos:
            archivo_seleccionado = st.selectbox("Selecciona un archivo:", st.session_state.archivos)

            # Mostrar el archivo seleccionado
            if archivo_seleccionado:
                #st.write(f"Has seleccionado: {archivo_seleccionado}")

                # Leer el contenido del archivo
                ruta_archivo = os.path.join(carpeta, archivo_seleccionado)

                # Determinar el tipo de archivo y leerlo
                if archivo_seleccionado.endswith('.txt'):
                    try:
                        with open(ruta_archivo, 'r', encoding='utf-8') as file:
                            contenido = file.read()
                            st.text_area("Contenido del archivo de texto:", contenido, height=300)
                    except UnicodeDecodeError:
                        st.error("Error al leer el archivo de texto. Intenta con otra codificación.")
                
                elif archivo_seleccionado.endswith('.xlsx') or archivo_seleccionado.endswith('.xls'):
                    try:
                        folder_path = ruta_archivo  # Cambia esto por la ruta real de tu carpeta
                        if os.name == 'nt':  # Si es Windows
                            os.startfile(folder_path)
                    except Exception as e:
                        st.error(f"Error al leer el archivo de Excel: {e}")
        else:
            pass
            #st.write("No hay archivos disponibles. Presiona el botón para cargar.")

    with col2:
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/digitalizacion.png"

        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=250)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #4CAF50; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                }
                
                .stButton>button:hover {
                    background-color: #45a049; /* Color de fondo al pasar el mouse */
                }
            </style>
        """, unsafe_allow_html=True)

        # Usar un botón para abrir la carpeta local
        if st.button("CAJAS DIGITALIZACION"):
            # Este código solo funciona en tu máquina local y abrirá la carpeta local especificada
            folder_path = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/CAJAS POR DIGITALIZAR/CAJAS POR DIGITALIZAR/Cajas_por_Digitalizar(08-11-2024).xlsx"  # Cambia esto por la ruta real de tu carpeta
            if os.name == 'nt':  # Si es Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # Si es Linux o macOS
                subprocess.Popen(['open', folder_path])  # En macOS
                # subprocess.Popen(['xdg-open', folder_path])  # En Linux

    with col3:
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/inventario.png"

        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=250)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #4CAF50; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                }
                
                .stButton>button:hover {
                    background-color: #45a049; /* Color de fondo al pasar el mouse */
                }
            </style>
        """, unsafe_allow_html=True)

        # Usar un botón para abrir la carpeta local
        if st.button("CAJAS INVENTARIO"):
            # Este código solo funciona en tu máquina local y abrirá la carpeta local especificada
            folder_path = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/TRASLADO REGIONALRES/REGIONALES.xlsx"  # Cambia esto por la ruta real de tu carpeta
            if os.name == 'nt':  # Si es Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # Si es Linux o macOS
                subprocess.Popen(['open', folder_path])  # En macOS
                # subprocess.Popen(['xdg-open', folder_path])  # En Linux

    
    with col4:
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/prueva3.png"

        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=220)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #4CAF50; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                }
                
                .stButton>button:hover {
                    background-color: #45a049; /* Color de fondo al pasar el mouse */
                }
            </style>
        """, unsafe_allow_html=True)

        # Usar un botón para abrir la carpeta local
        if st.button("CONTROL PRECINTOS"):
            # Este código solo funciona en tu máquina local y abrirá la carpeta local especificada
            folder_path = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/GUIAS_RECEPCION/CONTROL_DE_PRECINTOS.xlsx"  # Cambia esto por la ruta real de tu carpeta
            if os.name == 'nt':  # Si es Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # Si es Linux o macOS
                subprocess.Popen(['open', folder_path])  # En macOS
                # subprocess.Popen(['xdg-open', folder_path])  # En Linux

    st.markdown("<br><br>", unsafe_allow_html=True)

     # Crear columnas con tamaños personalizados (proporcionales)
    cl1, cl2, cl3, cl4 = st.columns(4)  # La segunda columna será más ancha
    

    # Colocar las imágenes en las columnas
    with cl1:  
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/Destuccion.png"

        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=250)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #9932CC; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                    }
                        
                .stButton>button:hover {
                    background-color: #9932CC; /* Color de fondo al pasar el mouse */
                    }
            </style>
        """, unsafe_allow_html=True)

        # Especifica la ruta de la carpeta que deseas listar
        carpeta = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/DESTRUCCION_BANCO_SOL"  # Cambia esto a la ruta de tu carpeta

        # Inicializa una variable de sesión para almacenar los archivos
        if 'archivos2' not in st.session_state:
            st.session_state.archivos2 = []

        # Función para listar archivos
        def listar_archivos():
            # Listar archivos en la carpeta
            st.session_state.archivos2 = os.listdir(carpeta)
            # Filtrar solo archivos (opcional)
            st.session_state.archivos2 = [archivo2 for archivo2 in st.session_state.archivos2 if os.path.isfile(os.path.join(carpeta, archivo2))]

        # Aplicar estilos CSS personalizados al botón

        # Botón para cargar archivos
        if st.button("DESTRUCCION DE CAJAS"):
            listar_archivos()
            #st.success("Archivos cargados.")

        # Mostrar los archivos en un selectbox si hay archivos disponibles
        if st.session_state.archivos2:
            archivo_seleccionado2 = st.selectbox("Selecciona un archivo:", st.session_state.archivos2)

            # Mostrar el archivo seleccionado
            if archivo_seleccionado2:
                #st.write(f"Has seleccionado: {archivo_seleccionado}")

                # Leer el contenido del archivo
                ruta_archivo = os.path.join(carpeta, archivo_seleccionado2)

                # Determinar el tipo de archivo y leerlo
                if archivo_seleccionado2.endswith('.txt'):
                    try:
                        with open(ruta_archivo, 'r', encoding='utf-8') as file:
                            contenido = file.read()
                            st.text_area("Contenido del archivo de texto:", contenido, height=300)
                    except UnicodeDecodeError:
                        st.error("Error al leer el archivo de texto. Intenta con otra codificación.")
                
                elif archivo_seleccionado2.endswith('.xlsx') or archivo_seleccionado2.endswith('.xls'):
                    try:
                        folder_path = ruta_archivo  # Cambia esto por la ruta real de tu carpeta
                        if os.name == 'nt':  # Si es Windows
                            os.startfile(folder_path)
                    except Exception as e:
                        st.error(f"Error al leer el archivo de Excel: {e}")
        else:
            pass
            #st.write("No hay archivos disponibles. Presiona el botón para cargar.")


    with cl2:
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/traslados.png"

        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=250)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #9932CC; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                    }
                        
                .stButton>button:hover {
                    background-color: #9932CC; /* Color de fondo al pasar el mouse */
                    }
            </style>
        """, unsafe_allow_html=True)

        # Especifica la ruta de la carpeta que deseas listar
        carpeta = "C:/Users/juan.ramos/Desktop/TRASLADOS - BNB"  # Cambia esto a la ruta de tu carpeta

        # Inicializa una variable de sesión para almacenar los archivos
        if 'archivos3' not in st.session_state:
            st.session_state.archivos3 = []

        # Función para listar archivos
        def listar_archivos():
            # Listar archivos en la carpeta
            st.session_state.archivos3 = os.listdir(carpeta)
            # Filtrar solo archivos (opcional)
            st.session_state.archivos3 = [archivo3 for archivo3 in st.session_state.archivos3 if os.path.isfile(os.path.join(carpeta, archivo3))]

        # Aplicar estilos CSS personalizados al botón


        # Botón para cargar archivos
        if st.button("TRASLADO BNB"):
            listar_archivos()
            #st.success("Archivos cargados.")

        # Mostrar los archivos en un selectbox si hay archivos disponibles
        if st.session_state.archivos3:
            archivo_seleccionado3 = st.selectbox("Selecciona un archivo:", st.session_state.archivos3)

            # Mostrar el archivo seleccionado
            if archivo_seleccionado3:
                #st.write(f"Has seleccionado: {archivo_seleccionado}")

                # Leer el contenido del archivo
                ruta_archivo = os.path.join(carpeta, archivo_seleccionado3)

                # Determinar el tipo de archivo y leerlo
                if archivo_seleccionado3.endswith('.txt'):
                    try:
                        with open(ruta_archivo, 'r', encoding='utf-8') as file:
                            contenido = file.read()
                            st.text_area("Contenido del archivo de texto:", contenido, height=300)
                    except UnicodeDecodeError:
                        st.error("Error al leer el archivo de texto. Intenta con otra codificación.")
                
                elif archivo_seleccionado3.endswith('.xlsx') or archivo_seleccionado3.endswith('.xls'):
                    try:
                        folder_path = ruta_archivo  # Cambia esto por la ruta real de tu carpeta
                        if os.name == 'nt':  # Si es Windows
                            os.startfile(folder_path)
                    except Exception as e:
                        st.error(f"Error al leer el archivo de Excel: {e}")
        else:
            pass
            #st.write("No hay archivos disponibles. Presiona el botón para cargar.")


    with cl3:
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/ruta.png"

        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=200)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    
                    background-color: #4CAF50; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                }
                
                .stButton>button:hover {
                    background-color: #45a049; /* Color de fondo al pasar el mouse */
                }
            </style>
        """, unsafe_allow_html=True)

        # Usar un botón para abrir la carpeta local
        if st.button("PROGRAMACION RUTAS"):
            # Este código solo funciona en tu máquina local y abrirá la carpeta local especificada
            folder_path = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/GUIAS_RECEPCION/RUTA-ALMACEN.xlsx"  # Cambia esto por la ruta real de tu carpeta
            if os.name == 'nt':  # Si es Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # Si es Linux o macOS
                subprocess.Popen(['open', folder_path])  # En macOS
                # subprocess.Popen(['xdg-open', folder_path])  # En Linux

    
    with cl4:
        # Ruta de la imagen
        image_url = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/plantilla.png"

        # Mostrar la imagen con un ancho específico
        st.image(image_url, width=270)  # Cambia 300 por el ancho que desees

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #9932CC; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                    }
                        
                .stButton>button:hover {
                    background-color: #9932CC; /* Color de fondo al pasar el mouse */
                    }
            </style>
        """, unsafe_allow_html=True)

        # Especifica la ruta de la carpeta que deseas listar
        carpeta = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\PLANTILLAS"  # Cambia esto a la ruta de tu carpeta

        # Inicializa una variable de sesión para almacenar los archivos
        if 'archivos4' not in st.session_state:
            st.session_state.archivos4 = []

        # Función para listar archivos
        def listar_archivos():
            # Listar archivos en la carpeta
            st.session_state.archivos4 = os.listdir(carpeta)
            # Filtrar solo archivos (opcional)
            st.session_state.archivos4 = [archivo4 for archivo4 in st.session_state.archivos4 if os.path.isfile(os.path.join(carpeta, archivo4))]

        # Aplicar estilos CSS personalizados al botón

        # Botón para cargar archivos
        if st.button("PLANTILLAS"):
            listar_archivos()
            #st.success("Archivos cargados.")

        # Mostrar los archivos en un selectbox si hay archivos disponibles
        if st.session_state.archivos4:
            archivo_seleccionado4 = st.selectbox("Selecciona un archivo:", st.session_state.archivos4)

            # Mostrar el archivo seleccionado
            if archivo_seleccionado4:
                #st.write(f"Has seleccionado: {archivo_seleccionado}")

                # Leer el contenido del archivo
                ruta_archivo = os.path.join(carpeta, archivo_seleccionado4)

                # Determinar el tipo de archivo y leerlo
                if archivo_seleccionado4.endswith('.txt'):
                    try:
                        with open(ruta_archivo, 'r', encoding='utf-8') as file:
                            contenido = file.read()
                            st.text_area("Contenido del archivo de texto:", contenido, height=300)
                    except UnicodeDecodeError:
                        st.error("Error al leer el archivo de texto. Intenta con otra codificación.")
                
                elif archivo_seleccionado4.endswith('.xlsx') or archivo_seleccionado4.endswith('.xls'):
                    try:
                        folder_path = ruta_archivo  # Cambia esto por la ruta real de tu carpeta
                        if os.name == 'nt':  # Si es Windows
                            os.startfile(folder_path)
                    except Exception as e:
                        st.error(f"Error al leer el archivo de Excel: {e}")
        else:
            pass
            #st.write("No hay archivos disponibles. Presiona el botón para cargar.")


    st.markdown("<br><br>", unsafe_allow_html=True)

    clum1, clum2, clum3, clum4 = st.columns(4)  # La segunda columna será más ancha
    

    # Colocar las imágenes en las columnas
    with clum1:  
        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #9932CC; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                    }
                        
                .stButton>button:hover {
                    background-color: #9932CC; /* Color de fondo al pasar el mouse */
                    }
            </style>
        """, unsafe_allow_html=True)

        # Especifica la ruta de la carpeta que deseas listar
        carpeta = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/REGULARIZACION_BPO"  # Cambia esto a la ruta de tu carpeta

        # Inicializa una variable de sesión para almacenar los archivos
        if 'archivos2' not in st.session_state:
            st.session_state.archivos2 = []

        # Función para listar archivos
        def listar_archivos():
            # Listar archivos en la carpeta
            st.session_state.archivos2 = os.listdir(carpeta)
            # Filtrar solo archivos (opcional)
            st.session_state.archivos2 = [archivo2 for archivo2 in st.session_state.archivos2 if os.path.isfile(os.path.join(carpeta, archivo2))]

        # Aplicar estilos CSS personalizados al botón


        # Botón para cargar archivos
        if st.button("REGULARIZACION FILES BPO"):
            listar_archivos()
            #st.success("Archivos cargados.")

        # Mostrar los archivos en un selectbox si hay archivos disponibles
        if st.session_state.archivos2:
            archivo_seleccionado2 = st.selectbox("Selecciona un archivo:", st.session_state.archivos2)

            # Mostrar el archivo seleccionado
            if archivo_seleccionado2:
                #st.write(f"Has seleccionado: {archivo_seleccionado}")

                # Leer el contenido del archivo
                ruta_archivo = os.path.join(carpeta, archivo_seleccionado2)

                # Determinar el tipo de archivo y leerlo
                if archivo_seleccionado2.endswith('.txt'):
                    try:
                        with open(ruta_archivo, 'r', encoding='utf-8') as file:
                            contenido = file.read()
                            st.text_area("Contenido del archivo de texto:", contenido, height=300)
                    except UnicodeDecodeError:
                        st.error("Error al leer el archivo de texto. Intenta con otra codificación.")
                
                elif archivo_seleccionado2.endswith('.xlsx') or archivo_seleccionado2.endswith('.xls'):
                    try:
                        folder_path = ruta_archivo  # Cambia esto por la ruta real de tu carpeta
                        if os.name == 'nt':  # Si es Windows
                            os.startfile(folder_path)
                    except Exception as e:
                        st.error(f"Error al leer el archivo de Excel: {e}")
        else:
            pass
            #st.write("No hay archivos disponibles. Presiona el botón para cargar.")

    with clum2:
        
        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #9932CC; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                    }
                        
                .stButton>button:hover {
                    background-color: #9932CC; /* Color de fondo al pasar el mouse */
                    }
            </style>
        """, unsafe_allow_html=True)

        # Especifica la ruta de la carpeta que deseas listar
        carpeta = "C:/Users/juan.ramos/Desktop/TRASLADOS - BNB"  # Cambia esto a la ruta de tu carpeta

        # Inicializa una variable de sesión para almacenar los archivos
        if 'archivos6' not in st.session_state:
            st.session_state.archivos6 = []

        # Función para listar archivos
        def listar_archivos():
            # Listar archivos en la carpeta
            st.session_state.archivos6 = os.listdir(carpeta)
            # Filtrar solo archivos (opcional)
            st.session_state.archivos6 = [archivo6 for archivo6 in st.session_state.archivos3 if os.path.isfile(os.path.join(carpeta, archivo6))]

        # Aplicar estilos CSS personalizados al botón

        # Botón para cargar archivos
        if st.button("FORMULARIOS DE VACACIONES"):
            listar_archivos()
            #st.success("Archivos cargados.")

        # Mostrar los archivos en un selectbox si hay archivos disponibles
        if st.session_state.archivos6:
            archivo_seleccionado6 = st.selectbox("Selecciona un archivo:", st.session_state.archivos6)

            # Mostrar el archivo seleccionado
            if archivo_seleccionado6:
                #st.write(f"Has seleccionado: {archivo_seleccionado}")

                # Leer el contenido del archivo
                ruta_archivo = os.path.join(carpeta, archivo_seleccionado6)

                # Determinar el tipo de archivo y leerlo
                if archivo_seleccionado6.endswith('.txt'):
                    try:
                        with open(ruta_archivo, 'r', encoding='utf-8') as file:
                            contenido = file.read()
                            st.text_area("Contenido del archivo de texto:", contenido, height=300)
                    except UnicodeDecodeError:
                        st.error("Error al leer el archivo de texto. Intenta con otra codificación.")
                
                elif archivo_seleccionado3.endswith('.xlsx') or archivo_seleccionado6.endswith('.xls'):
                    try:
                        folder_path = ruta_archivo  # Cambia esto por la ruta real de tu carpeta
                        if os.name == 'nt':  # Si es Windows
                            os.startfile(folder_path)
                    except Exception as e:
                        st.error(f"Error al leer el archivo de Excel: {e}")
        else:
            pass
            #st.write("No hay archivos disponibles. Presiona el botón para cargar.")


    with clum3:
    
        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    
                    background-color: #4CAF50; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                }
                
                .stButton>button:hover {
             um       background-color: #45a049; /* Color de fondo al pasar el mouse */
                }
            </style>
        """, unsafe_allow_html=True)

        # Usar un botón para abrir la carpeta local
        if st.button("FORMATO DE CONTROL DE TEMPERATURA"):
            # Este código solo funciona en tu máquina local y abrirá la carpeta local especificada
            folder_path = "C:/Users/juan.ramos/Desktop/Documentos JuanC/ARCHIVOS_POLYSITEMAS_JC/GUIAS_RECEPCION/RUTA-ALMACEN.xlsx"  # Cambia esto por la ruta real de tu carpeta
            if os.name == 'nt':  # Si es Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # Si es Linux o macOS
                subprocess.Popen(['open', folder_path])  # En macOS
                # subprocess.Popen(['xdg-open', folder_path])  # En Linux

    
    with clum4:

        # Aplicar estilos CSS personalizados al botón
        st.markdown("""
            <style>
                .stButton>button {
                    background-color: #9932CC; /* Color de fondo del botón (verde) */
                    color: white;              /* Color del texto */
                    font-size: 18px;           /* Tamaño de la fuente */
                    padding: 12px 24px;        /* Espaciado dentro del botón */
                    border-radius: 8px;        /* Bordes redondeados */
                    font-weight: bold;         /* Texto en negrita */
                    border: none;              /* Sin borde */
                    box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.2); /* Sombra sutil */
                    cursor: pointer;          /* Cursor en forma de mano */
                    transition: background-color 0.3s; /* Transición suave en hover */
                    }
                        
                .stButton>button:hover {
                    background-color: #9932CC; /* Color de fondo al pasar el mouse */
                    }
            </style>
        """, unsafe_allow_html=True)

        # Especifica la ruta de la carpeta que deseas listar
        carpeta = "C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\PLANTILLAS"  # Cambia esto a la ruta de tu carpeta

        # Inicializa una variable de sesión para almacenar los archivos
        if 'archivos4' not in st.session_state:
            st.session_state.archivos4 = []

        # Función para listar archivos
        def listar_archivos():
            # Listar archivos en la carpeta
            st.session_state.archivos4 = os.listdir(carpeta)
            # Filtrar solo archivos (opcional)
            st.session_state.archivos4 = [archivo4 for archivo4 in st.session_state.archivos4 if os.path.isfile(os.path.join(carpeta, archivo4))]

        # Aplicar estilos CSS personalizados al botón


        # Botón para cargar archivos
        if st.button("INFORMES ALMACEN"):
            listar_archivos()
            #st.success("Archivos cargados.")

        # Mostrar los archivos en un selectbox si hay archivos disponibles
        if st.session_state.archivos4:
            archivo_seleccionado4 = st.selectbox("Selecciona un archivo:", st.session_state.archivos4)

            # Mostrar el archivo seleccionado
            if archivo_seleccionado4:
                #st.write(f"Has seleccionado: {archivo_seleccionado}")

                # Leer el contenido del archivo
                ruta_archivo = os.path.join(carpeta, archivo_seleccionado4)

                # Determinar el tipo de archivo y leerlo
                if archivo_seleccionado4.endswith('.txt'):
                    try:
                        with open(ruta_archivo, 'r', encoding='utf-8') as file:
                            contenido = file.read()
                            st.text_area("Contenido del archivo de texto:", contenido, height=300)
                    except UnicodeDecodeError:
                        st.error("Error al leer el archivo de texto. Intenta con otra codificación.")
                
                elif archivo_seleccionado4.endswith('.xlsx') or archivo_seleccionado4.endswith('.xls'):
                    try:
                        folder_path = ruta_archivo  # Cambia esto por la ruta real de tu carpeta
                        if os.name == 'nt':  # Si es Windows
                            os.startfile(folder_path)
                    except Exception as e:
                        st.error(f"Error al leer el archivo de Excel: {e}")
        else:
            pass
            #st.write("No hay archivos disponibles. Presiona el botón para cargar.")  

elif proyectos == "DASHBOARD ALMACEN":

    #st.set_page_config(page_title="Dashboard_Almacen!!!", page_icon=":bar_chart:",layout="wide")

    st.title(" :bar_chart: DASHBOARD ALMACEN LA PAZ")
    st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

    fl = st.file_uploader(":file_folder: Upload a file",type=(["csv","txt","xlsx","xls"]))

    if fl is not None:
        filename = fl.name
        st.write(filename)
        df = pd.read_excel(filename)
    else:
        #os.chdir(r"C:/Users/Juan_C/Desktop/Prueba_Python")
        #df = pd.read_excel("Adidas.xlsx")

        #excel_file = 'C:/Users/Juan_C/Desktop/Prueba_Python/ALMACEN.xlsx'
        excel_file = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/ALMACEN.xlsx'
        sheet_name = 'INGRESO CAJAS POR CLIENTE '
        df = pd.read_excel(excel_file, sheet_name=sheet_name)

        excel_file2 = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/ALMACEN.xlsx'
        sheet_name = 'CAJAS DESTRUIDAS'
        Destruccion = pd.read_excel(excel_file2, sheet_name=sheet_name)

        excel_file2 = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/ALMACEN.xlsx'
        sheet_name = 'TOTAL INGRESO DE CAJAS'
        Ingreso_Cajas = pd.read_excel(excel_file2, sheet_name=sheet_name)

        excel_file3 = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/Control de Cajas Poly.xlsx'
        sheet_name = 'Stock de Cajas'
        Stock_Cajas = pd.read_excel(excel_file3, sheet_name=sheet_name)

        excel_file4 = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/Control de Cajas Poly.xlsx'
        sheet_name = 'Stock Total Cintillos'
        Stock_Cintillos = pd.read_excel(excel_file4, sheet_name=sheet_name)

    Total_Cajas = Ingreso_Cajas['TOTAL INGRESO DE CAJAS '].sum()
    Stock_Cajas_Almacen = Stock_Cajas['TOTAL STOCK ACTUAL'].sum()
    Stock_Cintillos_Almacen = Stock_Cintillos['TOTAL STOCK ACTUAL CINTILLOS'].sum()

    cl1, cl2, cl3 = st.columns((3))
    with cl1:
        # Mostrar el total de ingresos centrado en una caja
        st.markdown(
            f"""
            <div style='
                border: 5px solid black; 
                padding: 5px; 
                border-radius: 2px; 
                text-align: center; 
                background-color: #060606;'>
                <h3 style='color: white;'>TOTAL INGRESO DE CAJAS POLYSISTEMAS LA PAZ</h3>
                <h1 style='color: white;'>{Total_Cajas:,.2f}</h1>
            </div>
            """, 
            unsafe_allow_html=True
        )

    with cl2:
        # Mostrar el total de ingresos centrado en una caja
        st.markdown(
            f"""
            <div style='
                border: 5px solid black; 
                padding: 5px; 
                border-radius: 2px; 
                text-align: center; 
                background-color: #060606;'>
                <h3 style='color: white;'>TOTAL STOCK DE CAJAS PLANAS POLYSISTEMAS LA PAZ</h3>
                <h1 style='color: white;'>{Stock_Cajas_Almacen:,.2f}</h1>
            </div>
            """, 
            unsafe_allow_html=True
        )

    with cl3:
        # Mostrar el total de ingresos centrado en una caja
        st.markdown(
            f"""
            <div style='
                border: 5px solid black; 
                padding: 5px; 
                border-radius: 2px; 
                text-align: center; 
                background-color: #060606;'>
                <h3 style='color: white;'>TOTAL STOCK CINTILLOS POLYSISTEMAS LA PAZ</h3>
                <h1 style='color: white;'>{Stock_Cintillos_Almacen:,.2f}</h1>
            </div>
            """, 
            unsafe_allow_html=True
        )

    col1, col2 = st.columns((2))
    df["FECHA  DE INGRESO"] = pd.to_datetime(df["FECHA  DE INGRESO"])

    # Getting the min and max date 
    startDate = pd.to_datetime(df["FECHA  DE INGRESO"]).min()
    endDate = pd.to_datetime(df["FECHA  DE INGRESO"]).max()

    with col1:
        date1 = pd.to_datetime(st.date_input("Fecha Inicio", startDate))

    with col2:
        date2 = pd.to_datetime(st.date_input("Fecha Fin", endDate))

    df = df[(df["FECHA  DE INGRESO"] >= date1) & (df["FECHA  DE INGRESO"] <= date2)].copy()


    st.sidebar.header("FILTARAR: ")
    # Create for Region
    cliente = st.sidebar.multiselect("PROYECTOS", df["CLIENTE"].unique())
    if not cliente:
        df2 = df.copy()
    else:
        df2 = df[df["CLIENTE"].isin(cliente)]

    # Create for State
    lugar_recojo = st.sidebar.multiselect("Lugar de Ingreso ", df2["LUGAR DE RECOJO"].unique())
    if not lugar_recojo:
        df3 = df2.copy()
    else:
        df3 = df2[df2["LUGAR DE RECOJO"].isin(lugar_recojo)]


    # Create for City
    agencia = st.sidebar.multiselect("Agencia",df3["AGENCIA"].unique())

    # Filter the data based on Region, State and City

    if not cliente and not lugar_recojo and not agencia:
        filtered_df = df
    elif not lugar_recojo and not agencia:
        filtered_df = df[df["CLIENTE"].isin(cliente)]
    elif not cliente and not agencia:
        filtered_df = df[df["LUGAR DE RECOJO"].isin(lugar_recojo)]
    elif lugar_recojo and agencia:
        filtered_df = df3[df["LUGAR DE RECOJO"].isin(lugar_recojo) & df3["AGENCIA"].isin(agencia)]
    elif cliente and agencia:
        filtered_df = df3[df["CLIENTE"].isin(cliente) & df3["AGENCIA"].isin(agencia)]
    elif cliente and lugar_recojo:
        filtered_df = df3[df["CLIENTE"].isin(cliente) & df3["LUGAR DE RECOJO"].isin(lugar_recojo)]
    elif agencia:
        filtered_df = df3[df3["AGENCIA"].isin(agencia)]
    else:
        filtered_df = df3[df3["CLIENTE"].isin(cliente) & df3["LUGAR DE RECOJO"].isin(lugar_recojo) & df3["AGENCIA"].isin(agencia)]

    cliente_df = filtered_df.groupby(by = ["CLIENTE"], as_index = False)["CANTIDAD TOTAL"].sum()

    with col1:

        tipo_grafico = st.selectbox(
            "Selecciona el tipo de grafico",
            options=["Barras", "Diagrama de Dispercion"]
        )

        if tipo_grafico == "Barras":

            st.subheader("CANTIDAD DE CAJAS POR CLIENTE")
            fig = px.bar(cliente_df, x = "CLIENTE", y = "CANTIDAD TOTAL", text = ['{:,.2f}'.format(x) for x in cliente_df["CANTIDAD TOTAL"]],
                        template = "seaborn")
            st.plotly_chart(fig,use_container_width=True, height = 200)

        elif tipo_grafico == "Diagrama de Dispercion":
            data1 = px.scatter(cliente_df, x = "CLIENTE", y = "CANTIDAD TOTAL", size = "CANTIDAD TOTAL")
            data1['layout'].update(title="RELACION ENTRE CANTIDAD y CLIENTE DIAGRAMA DE DISPERCION.",
                        titlefont = dict(size=20),xaxis = dict(title="CLIENTES",titlefont=dict(size=19)),
                        yaxis = dict(title = "CANTIDAD CAJAS", titlefont = dict(size=19)))
            st.plotly_chart(data1,use_container_width=True)

    with col2:

        tipo_grafico = st.selectbox(
        "Selecciona el tipo de grafico",
        options=["Torta", "Diagrama de Dispercion"]
        )

        if tipo_grafico == "Torta":
            st.subheader("CANTIDAD DE CAJAS POR AGENCIA")
            fig = px.pie(filtered_df, values="CANTIDAD TOTAL", names="LUGAR DE RECOJO", hole=0.5)
            # Mostrar solo label y valor (sin porcentaje)
            fig.update_traces(textinfo='label+value', textposition='outside', hoverinfo='label+value')
            st.plotly_chart(fig, use_container_width=True)

        elif tipo_grafico == "Diagrama de Dispercion":
            data1 = px.scatter(filtered_df, x="LUGAR DE RECOJO", y="CANTIDAD TOTAL", size="CANTIDAD TOTAL")
            data1['layout'].update(title="RELACION ENTRE CANTIDAD y LUGAR DE RECOJO DIAGRAMA DE DISPERCION.",
                               titlefont=dict(size=20), xaxis=dict(title="LUGAR DE RECOJO", titlefont=dict(size=19)),
                               yaxis=dict(title="CANTIDAD CAJAS", titlefont=dict(size=19)))
            st.plotly_chart(data1, use_container_width=True)


    cl1, cl2 = st.columns((2))
    with cl1:
        with st.expander("Ver Datos por Cliente"):
            st.write(cliente_df.style.background_gradient(cmap="Blues"))
            csv = cliente_df.to_csv(index = False).encode('utf-8')
            st.download_button("Download Data", data = csv, file_name = "Cantidad_Cajas_Clientes.csv", mime = "text/xlsx", help = 'Click here to download the data as a EXCEL file')

    with cl2:
        with st.expander("Ver Datos por Agencia"):
            region = filtered_df.groupby(by = "LUGAR DE RECOJO", as_index = False)["CANTIDAD TOTAL"].sum()
            st.write(region.style.background_gradient(cmap="Oranges"))
            csv = region.to_csv(index = False).encode('utf-8')
            st.download_button("Download Data", data = csv, file_name = "Cantidad_Cajas_Agencia.csv", mime = "text/csv",
                            help = 'Click here to download the data as a CSV file')
            

    chart1, chart2 = st.columns((2))
    with chart1:
        st.subheader('CAJAS DESTRUIDAS - SALIDA PERMANENTE')
        fig = px.pie(Destruccion, values="CANTIDAD", names="DESCRIPCION", template="plotly_dark")
        # Mostrar solo label y valor (sin porcentaje)
        fig.update_traces(textinfo='label+value', textposition='inside', hoverinfo='label+value')
        st.plotly_chart(fig, use_container_width=True)

    with chart2:
        # Crear un diagrama de dispersión
        data1 = px.scatter(Destruccion, x="FECHA ACTUAL", y="CANTIDAD", size="CANTIDAD")
        data1['layout'].update(title="RELACION ENTRE CANTIDAD y FECHA DE SALIDA DIAGRAMA DE DISPERCION.",
                           titlefont=dict(size=20), xaxis=dict(title="FECHA DE SALIDA", titlefont=dict(size=19)),
                           yaxis=dict(title="CANTIDAD DE CAJAS", titlefont=dict(size=19)))
        st.plotly_chart(data1, use_container_width=True)

    cls1, cls2 =st.columns((2))
            
    st.subheader("MAPA DE CAJAS DESTRUIDAS Y DE SALIDA PERMANENTE")
    fig5 = px.treemap(
        Destruccion, 
        path = ["DESCRIPCION","FECHA ACTUAL", "CANTIDAD", "CLIENTE"], 
        values = "CANTIDAD",
        #hover_data = ["FECHA ACTUAL"],
        color = "CANTIDAD",
        color_continuous_scale = 'Blues',
        #title = 'Mapa de Cajas Destruidas'
    )
    fig5.update_layout(width = 1300, height = 350)
    st.plotly_chart(fig5, use_container_width=True)

    with cls1:
        with st.expander("Ver datos Cajas Destruccion"):
            region = Destruccion.groupby(by = ['CLIENTE', 'DESCRIPCION', 'FECHA ACTUAL'], as_index = False)["CANTIDAD"].sum()
            st.write(region.style.background_gradient(cmap="Blues"))
            csv = region.to_csv(index = False).encode('utf-8')
            st.download_button("Download Data", data = csv, file_name = "Destruccion_de_Cajas.csv", mime = "text/csv",
                            help = 'Click here to download the data as a CSV file')
            
    with cls2:
        with st.expander("Ver datos Cajas Destruccion"):
            region = Destruccion.groupby(by = ['CLIENTE', 'DESCRIPCION', 'FECHA ACTUAL'], as_index = False)["CANTIDAD"].sum()
            st.write(region.style.background_gradient(cmap="Greens"))
            csv2 = region.to_csv(index = False).encode('utf-8')
            st.download_button("Download Data", data = csv2, file_name = "Destruccion_de_Cajas2.csv", mime = "text/csv",
                            help = 'Click here to download the data as a CSV file')

elif proyectos == "DASHBOARD DESPACHO":

    st.title(" :bar_chart: DASHBOARD DESPACHO LA PAZ")
    st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

    fl = st.file_uploader(":file_folder: Upload a file",type=(["csv","txt","xlsx","xls"]))

    if fl is not None:
        filename = fl.name
        st.write(filename)
        df = pd.read_excel(filename)
    else:
        #os.chdir(r"C:/Users/Juan_C/Desktop/Prueba_Python")
        #df = pd.read_excel("Adidas.xlsx")

        #excel_file = 'C:/Users/Juan_C/Desktop/Prueba_Python/ALMACEN.xlsx'
        excel_file = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/Extraccion_Despacho.xlsx'
        sheet_name = 'Extraccion Despacho'
        dfd = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        # Eliminar filas que están completamente vacías
        dfE = dfd.dropna(how='all')

        # llenamos con Ceros todos los Nan
        df = dfE.fillna(0)

        #st.dataframe(df)

        excel_file = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN\Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/DOC. PROCESO DEVOLUCION.xlsx'
        sheet_name = 'TOTAL DOC. POR DEVOLVER'
        Total_Devolucion = pd.read_excel(excel_file, sheet_name=sheet_name)

        Total_Devolucion_BSol = Total_Devolucion['TOTAL DOC. POR DEVOLVER BSOL'].sum()
        Total_Devolucion_BNB = Total_Devolucion['TOTAL DOC. POR DEVOLVER BNB'].sum()
        Total_Devolucion_BFIE = Total_Devolucion['TOTAL DOC. POR DEVOLVER BFIE'].sum()

    st.title("EXTRACCION Y DEVOLUCION DE DOCUMENTOS ALMACEN A LA PAZ ")
    st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

    cl1, cl2, cl3 = st.columns((3))
    with cl1:
        # Mostrar el total de ingresos centrado en una caja
        st.markdown(
            f"""
            <div style='
                border: 5px solid black; 
                padding: 1px; 
                border-radius: 2px; 
                text-align: center; 
                background-color: #060606;'>
                <h3 style='color: white;'>TOTAL DOC. POR DEVOLVER BSOL POLYSISTEMAS LA PAZ</h3>
                <h1 style='color: white;'>{Total_Devolucion_BSol:,.2f}</h1>
            </div>
            """, 
            unsafe_allow_html=True
        )

    with cl2:
        # Mostrar el total de ingresos centrado en una caja
        st.markdown(
            f"""
            <div style='
                border: 5px solid black; 
                padding: 1px; 
                border-radius: 2px; 
                text-align: center; 
                background-color: #060606;'>
                <h3 style='color: white;'>TOTAL DOC. POR DEVOLVER BNB POLYSISTEMAS LA PAZ</h3>
                <h1 style='color: white;'>{Total_Devolucion_BNB:,.2f}</h1>
            </div>
            """, 
            unsafe_allow_html=True
        )

    with cl3:
        # Mostrar el total de ingresos centrado en una caja
        st.markdown(
            f"""
            <div style='
                border: 5px solid black; 
                padding: 1px; 
                border-radius: 2px; 
                text-align: center; 
                background-color: #060606;'>
                <h3 style='color: white;'>TOTAL DOC. POR DEVOLVER BFIE POLYSISTEMAS LA PAZ</h3>
                <h1 style='color: white;'>{Total_Devolucion_BFIE:,.2f}</h1>
            </div>
            """, 
            unsafe_allow_html=True
        )


    col1, col2 = st.columns((2))
    df["FECHA"] = pd.to_datetime(df["FECHA"])

    # Obtener min y max solo de fechas válidas
    startDate = pd.to_datetime(df["FECHA"]).min()
    endDate = pd.to_datetime(df["FECHA"]).max()

    with col1:
        date1 = pd.to_datetime(st.date_input("Fecha Inicio", startDate))

    with col2:
        date2 = pd.to_datetime(st.date_input("Fecha Fin", endDate))

    df = df[(df["FECHA"] >= date1) & (df["FECHA"] <= date2)].copy()

    # Logo 
    #st.sidebar.image("C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Dashboard_Poly/Poly_logo.png", use_column_width=True, caption="DASHBOARD ALMACEN LA PAZ")

    st.sidebar.header("FILTARAR: ")
    # Create for Region
    cliente = st.sidebar.multiselect("PROYECTOS", df["CLIENTE"].unique())
    if not cliente:
        df2 = df.copy()
    else:
        df2 = df[df["CLIENTE"].isin(cliente)]

    # Create for State
    Tipo_Extraccion = st.sidebar.multiselect("Tipo Extraccion ", df2["TIPO DE EXTRACCION"].unique())
    if not Tipo_Extraccion:
        df3 = df2.copy()
    else:
        df3 = df2[df2["TIPO DE EXTRACCION"].isin(Tipo_Extraccion)]


    # Create for City
    Tipo_Servicio = st.sidebar.multiselect("Tipo de Servicio",df3["TIPO DE SERVICIO"].unique())

    # Filter the data based on Region, State and City

    if not cliente and not Tipo_Extraccion and not Tipo_Servicio:
        filtered_df = df
    elif not Tipo_Extraccion and not Tipo_Servicio:
        filtered_df = df[df["CLIENTE"].isin(cliente)]
    elif not cliente and not Tipo_Servicio:
        filtered_df = df[df["TIPO DE EXTRACCION"].isin(Tipo_Extraccion)]
    elif Tipo_Extraccion and Tipo_Servicio:
        filtered_df = df3[df["TIPO DE EXTRACCION"].isin(Tipo_Extraccion) & df3["TIPO DE SERVICIO"].isin(Tipo_Servicio)]
    elif cliente and Tipo_Servicio:
        filtered_df = df3[df["CLIENTE"].isin(cliente) & df3["TIPO DE SERVICIO"].isin(Tipo_Servicio)]
    elif cliente and Tipo_Extraccion:
        filtered_df = df3[df["CLIENTE"].isin(cliente) & df3["TIPO DE EXTRACCION"].isin(Tipo_Extraccion)]
    elif Tipo_Servicio:
        filtered_df = df3[df3["TIPO DE SERVICIO"].isin(Tipo_Servicio)]
    else:
        filtered_df = df3[df3["CLIENTE"].isin(cliente) & df3["TIPO DE EXTRACCION"].isin(Tipo_Extraccion) & df3["TIPO DE SERVICIO"].isin(Tipo_Servicio)]

    cliente_df = filtered_df.groupby(by = ["CLIENTE"], as_index = False)["CANTIDAD DE ITEMS"].sum()

    with col1:

        tipo_grafico = st.selectbox(
            "Selecciona el tipo de grafico",
            options=["Barras", "Diagrama de Dispercion"]
        )

        if tipo_grafico == "Barras":

            st.subheader("EXTRACCION DE DOC. POR CLIENTE")
            fig = px.bar(cliente_df, x = "CLIENTE", y = "CANTIDAD DE ITEMS", text = ['{:,.2f}'.format(x) for x in cliente_df["CANTIDAD DE ITEMS"]],
                        template = "seaborn")
            st.plotly_chart(fig,use_container_width=True, height = 200)

        elif tipo_grafico == "Diagrama de Dispercion":
            data1 = px.scatter(cliente_df, x = "CLIENTE", y = "CANTIDAD DE ITEMS", size = "CANTIDAD DE ITEMS")
            data1['layout'].update(title="RELACION ENTRE CLIENTE Y CANTIDAD DE EXTRACCION DIAGRAMA DE DISPERCION.",
                        titlefont = dict(size=20),xaxis = dict(title="CLIENTES",titlefont=dict(size=19)),
                        yaxis = dict(title = "CANTIDAD DE ITEMS", titlefont = dict(size=19)))
            st.plotly_chart(data1,use_container_width=True)

    with col2:

        tipo_grafico = st.selectbox(
            "Selecciona el tipo de grafico",
            options=["Torta", "Diagrama de Dispercion", "Barras"]
        )

        if tipo_grafico == "Torta":
            st.subheader("CANTIDAD DE ITEMS POR TIPO DE EXTRACCION")
            fig = px.pie(filtered_df, values = "CANTIDAD DE ITEMS", names = "TIPO DE EXTRACCION", hole = 0.5)
            fig.update_traces(text = filtered_df["TIPO DE EXTRACCION"], textposition = "outside")
            st.plotly_chart(fig,use_container_width=True)

        elif tipo_grafico == "Diagrama de Dispercion":

            st.subheader("RELACION ENTRE CANTIDAD y TIPO DE EXTRACCION DIAGRAMA DE DISPERCION")
            #data1 = px.scatter(filtered_df, x = "FECHA", y = "CANTIDAD DE ITEMS", size = "CANTIDAD DE ITEMS")
            #data1['layout'].update(title="RELACION ENTRE CANTIDAD y TIPO DE EXTRACCION DIAGRAMA DE DISPERCION.",
            #               titlefont = dict(size=20),xaxis = dict(title="TIPO DE EXTRACCION",titlefont=dict(size=19)),
            #               yaxis = dict(title = "CANTIDAD DE ITEMS", titlefont = dict(size=19)))
            #st.plotly_chart(data1,use_container_width=True)
            scatter = alt.Chart(filtered_df).mark_circle(size=60).encode(
                x = 'FECHA',
                y = 'CANTIDAD DE ITEMS',
                tooltip = ['TIPO DE EXTRACCION', 'CANTIDAD DE ITEMS']
            ).interactive()

            st.altair_chart(scatter, use_container_width=True)

        elif tipo_grafico == "Barras":
            st.subheader("CANTIDAD DE ITEMS POR TIPO DE EXTRACCION")
            fig = go.Figure(data=[
                go.Bar(
                x=filtered_df['TIPO DE EXTRACCION'],
                y=filtered_df['CANTIDAD DE ITEMS'],
                marker=dict(line=dict(width=0.5), color= 'lightblue')  # Ajusta el grosor del borde
                )
            ])

            # Actualizar el layout para barras más estilizadas
            fig.update_traces(marker_line_width=8)  # Grosor del borde de la barra
            fig.update_layout(bargap=0.5)  # Separación entre barras

            # Mostrar gráfico en Streamlit
            st.plotly_chart(fig,use_container_width=True)



    cl1, cl2 = st.columns((2))
    with cl1:
        with st.expander("Ver Datos por Cliente"):
            st.write(cliente_df.style.background_gradient(cmap="Blues"))
            csv = cliente_df.to_csv(index = False).encode('utf-8')
            st.download_button("Download Data", data = csv, file_name = "Cantidad_Extraccion_Clientes.csv", mime = "text/xlsx", help = 'Click here to download the data as a EXCEL file')

    with cl2:
        with st.expander("Ver Datos por Extraccion"):
            region = filtered_df.groupby(by = ["TIPO DE EXTRACCION"], as_index = False)["CANTIDAD DE ITEMS"].sum()
            st.write(region.style.background_gradient(cmap="Oranges"))
            csv = region.to_csv(index = False).encode('utf-8')
            st.download_button("Download Data", data = csv, file_name = "Cantidad_Items_Extraidos.csv", mime = "text/csv",
                            help = 'Click here to download the data as a CSV file')
            

    # Create a treem based on Region, category, sub-Category
    st.subheader("MAPA DE EXTRACCION DE DOCUMENTOS")
    fig = px.treemap(filtered_df, path = ["TIPO DE EXTRACCION","CLIENTE"], values = "CANTIDAD DE ITEMS", hover_data = ["CANTIDAD DE ITEMS"],
                    color = "CANTIDAD DE ITEMS")
    fig.update_layout(width = 800, height = 650)
    st.plotly_chart(fig, use_container_width=True)


    st.title("INGRESO DE CAJAS POR REGIONAL ALMACEN LA PAZ")
    st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

    excel_file2 = 'C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Dashboard_Poly/DASHBOARD_ALMACEN_LA_PAZ/ENTRGA_DE_CAJAS_POLY.xlsx'
    sheet_name = 'Ingreso_Cajas'
    df2 = pd.read_excel(excel_file2, sheet_name=sheet_name)


    col1, col2 = st.columns((2))
    df2["FECHA DE INGRESO"] = pd.to_datetime(df2["FECHA DE INGRESO"], errors='coerce')

    # Filtrar fechas válidas
    valid_dates2 = df2["FECHA DE INGRESO"].dropna()

    # Obtener min y max solo de fechas válidas
    startDate = valid_dates2.min()
    endDate = valid_dates2.max()

    with col1:
        #date1 = pd.to_datetime(st.date_input("Fecha Inicio", startDate))
        date1 = pd.to_datetime(st.date_input("Fecha Inicio", startDate, key="fecha_inicio"))

    with col2:
        #date2 = pd.to_datetime(st.date_input("Fecha Fin", endDate))
        date2 = pd.to_datetime(st.date_input("Fecha Fin", endDate, key="fecha_fin"))

    df2 = df2[(df2["FECHA DE INGRESO"] >= date1) & (df2["FECHA DE INGRESO"] <= date2)].copy()


    colu_1, colu_2 = st.columns((2))

    with colu_1:

        tipo_grafico2 = st.selectbox(
            "Selecciona el tipo de grafico",
            options=["Barras", "Torta"]
        )

        if tipo_grafico2 == "Barras":

            st.subheader("CANTIDAD DE INGRESO DE CAJAS POR REGIONAL")
            fig = px.bar(df2, x = "REGIONAL", y = ["CANTIDAD"], text = ['{:,.2f}'.format(x) for x in df2["CANTIDAD"]],
                        template = "seaborn")
            st.plotly_chart(fig,use_container_width=True, height = 200)

        if tipo_grafico2 == "Torta":
            st.subheader("CANTIDAD DE INGRESO DE CAJAS POR CLIENTE")
            fig3 = px.pie(df2, values = "CANTIDAD", names = "CLIENTE", hole = 0.5)
            fig3.update_traces(text = df2["CANTIDAD"], textposition = "outside")
            st.plotly_chart(fig3,use_container_width=True)

    with colu_2:

        tipo_grafico = st.selectbox(
            "Selecciona el tipo de grafico",
            options=["Torta", "Diagrama de Dispercion"]
        )

        st.subheader("CANTIDAD DE INGRESO DE CAJAS POR AREA")
        # Crear dos botones para cambiar de gráfico
        grafico_seleccionado = st.radio("AREAS:", ["DIGITALIZACION", "INVENTARIO B_SOL", "INVENTARIO BNB"])

        if tipo_grafico == "Torta":
            if grafico_seleccionado == "DIGITALIZACION":
                
                fig = px.pie(df2, values = "DIGITALIZACION", names = "REGIONAL", hole = 0.5)
                fig.update_traces(text = df2["REGIONAL"], textposition = "outside")
                st.plotly_chart(fig,use_container_width=True)

            elif grafico_seleccionado == "INVENTARIO B_SOL":
                
                fig = px.pie(df2, values = "INVENTARIO B_SOL", names = "REGIONAL", hole = 0.5)
                fig.update_traces(text = df2["REGIONAL"], textposition = "outside")
                st.plotly_chart(fig,use_container_width=True)

            elif grafico_seleccionado == "INVENTARIO BNB":
                
                fig = px.pie(df2, values = "INVENTARIO BNB", names = "REGIONAL", hole = 0.5)
                fig.update_traces(text = df2["REGIONAL"], textposition = "outside")
                st.plotly_chart(fig,use_container_width=True)

        elif tipo_grafico == "Diagrama de Dispercion":
            if grafico_seleccionado == "DIGITALIZACION":
                scatter = alt.Chart(df2).mark_circle(size=60).encode(
                    x = 'FECHA DE INGRESO',
                    y = 'DIGITALIZACION',
                    tooltip = ['REGIONAL', 'DIGITALIZACION']
                ).interactive()

                st.altair_chart(scatter, use_container_width=True)
            
            elif grafico_seleccionado == "INVENTARIO B_SOL":
                scatter = alt.Chart(df2).mark_circle(size=60).encode(
                    x = 'FECHA DE INGRESO',
                    y = 'INVENTARIO B_SOL',
                    tooltip = ['REGIONAL', 'INVENTARIO B_SOL']
                ).interactive()

                st.altair_chart(scatter, use_container_width=True)

            elif grafico_seleccionado == "INVENTARIO BNB":
                scatter = alt.Chart(df2).mark_circle(size=60).encode(
                    x = 'FECHA DE INGRESO',
                    y = 'INVENTARIO BNB',
                    tooltip = ['REGIONAL', 'INVENTARIO BNB']
                ).interactive()

                st.altair_chart(scatter, use_container_width=True)

elif proyectos == "FILTRADOR ALMACEN":
    st.markdown("<h1 style='text-align: center;'>FILTRADO DE DOCUMENTOS Y CAJAS ALMACEN</h1>", unsafe_allow_html=True)

    opciones = ["SELECCIONA UNA OPCION", "FILTRADO FILES", "FILTRADO TOMOS", "FILTRADO DE CAJAS"]

    seleccion = st.selectbox("Selecciona una opcion del menu: ", opciones)


    if seleccion == "SELECCIONA UNA OPCION":   
        mensaje_markdown = """
        ### BIENVENIDO !! FILTRADO DE DOCUMENTOS ALMACEN

        **Desarrollado por Juan Carlos Ramos Chura**
        """
        st.markdown(mensaje_markdown)

    elif seleccion == "FILTRADO FILES":


        col1, col2 = st.columns(2)

        with col1:
            st.markdown("<h2 style='text-align: center;'>CARGAR PLANILLA DE EXCEL PARA FILTRAR FILES</h2>", unsafe_allow_html=True)

            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])


            if uploaded_file is not None:
            
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)
                
                ruta = st.text_input("Introduce la ruta de la carpeta para guardar los Archivos: Obligatorio", "C:\\")
        
                st.write('FILTRADO POR NIVELES:')

                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES","Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])
                
                Nivel = st.selectbox("Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":

                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                        file_path =  save_dir / f"{file_name}.xlsx"

                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "2":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )                           

                if Nivel == "3":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")


                        save_dir = Path(ruta)

                        file_path =  save_dir / f"{file_name}.xlsx"

                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "4":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )        

                if Nivel == "5":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")


                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "6":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
    
                # Mostrar un mensaje
                st.write('FILTRADO POR LOCACIONES:')

                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":
                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    st.dataframe(Ordenar)

                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)
                        
                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones
            #st.write("Sube varios archivos Excel para combinarlos en uno solo.")
            st.markdown("<h2 style='text-align: center;'>SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar múltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Función para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Files.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")


            # Título de la aplicación
            st.markdown("<h2 style='text-align: center;'>ELIMINAR ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Especificar la ruta de la carpeta donde están los archivos Excel
            folder_path = st.text_input("Introduce la ruta de la carpeta done Guardo Los Archivos Filtrados", "C:\\")

            # Comprobar si la ruta es válida y es una carpeta
            if folder_path and os.path.exists(folder_path) and os.path.isdir(folder_path):
                # Listar todos los archivos Excel en la carpeta
                excel_files = glob.glob(os.path.join(folder_path, "*.xlsx")) + glob.glob(os.path.join(folder_path, "*.xls"))

                # Mostrar la cantidad de archivos Excel encontrados
                st.write(f"Se encontraron {len(excel_files)} archivos Excel en la carpeta.")

                # Si hay archivos Excel, proporcionar la opción de eliminarlos
                if excel_files:
                    # Botón de confirmación para eliminar todos los archivos
                    if st.button("Eliminar todos los archivos Excel"):
                        try:
                            # Eliminar cada archivo encontrado
                            for file in excel_files:
                                os.remove(file)
                            st.success(f"Se eliminaron {len(excel_files)} archivos Excel de la carpeta.")
                        except Exception as e:
                            st.error(f"Error al eliminar archivos: {e}")
                else:
                    st.write("No se encontraron archivos Excel en la carpeta especificada.")
            else:
                st.write("Introduce una ruta válida para la carpeta.")

            # ---------------------------------------------------------------------------------------------------------

    elif seleccion == "FILTRADO TOMOS":

        col1, col2 = st.columns(2)
        with col1:
            # Titulo de Aplicacion
        
            st.markdown("<h2 style='text-align: center;'>CARGAR PLANILLA DE EXCEL PARA FILTRAR TOMOS</h2>", unsafe_allow_html=True)

        
            #Cargar el archivo de excel 
            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])

        
            if uploaded_file is not None:
                # Leer el archivo Excel usando Pandas
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                # Elimoinar Columnas
                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)

                # Definimos una ruta para guardar nuestros archivos
                ruta = st.text_input("Introduce la ruta de la carpeta para guardar los Archivos: Obligatorio", "C:\\")
                
            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])

                Nivel = st.selectbox("Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)
                        
                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "2":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "3":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "4":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "5":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "6":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":
                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)

                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":

                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones
            st.markdown("<h2 style='text-align: center;'>SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar múltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Función para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Tomos.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")

            # Título de la aplicación
        
            st.markdown("<h2 style='text-align: center;'>ELIMINAR ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Especificar la ruta de la carpeta donde están los archivos Excel
            folder_path = st.text_input("Introduce la ruta de la carpeta done Guardo Los Archivos Filtrados", "C:\\")

            # Comprobar si la ruta es válida y es una carpeta
            if folder_path and os.path.exists(folder_path) and os.path.isdir(folder_path):
                # Listar todos los archivos Excel en la carpeta
                excel_files = glob.glob(os.path.join(folder_path, "*.xlsx")) + glob.glob(os.path.join(folder_path, "*.xls"))

                # Mostrar la cantidad de archivos Excel encontrados
                st.write(f"Se encontraron {len(excel_files)} archivos Excel en la carpeta.")

                # Si hay archivos Excel, proporcionar la opción de eliminarlos
                if excel_files:
                    # Botón de confirmación para eliminar todos los archivos
                    if st.button("Eliminar todos los archivos Excel"):
                        try:
                            # Eliminar cada archivo encontrado
                            for file in excel_files:
                                os.remove(file)
                            st.success(f"Se eliminaron {len(excel_files)} archivos Excel de la carpeta.")
                        except Exception as e:
                            st.error(f"Error al eliminar archivos: {e}")
                else:
                    st.write("No se encontraron archivos Excel en la carpeta especificada.")
            else:
                st.write("Introduce una ruta válida para la carpeta.")

            # ---------------------------------------------------------------------------------------------------------


    elif seleccion == "FILTRADO DE CAJAS":

        col1, col2 = st.columns(2)
        with col1:
            # Titulo de Aplicacion
        
            st.markdown("<h2 style='text-align: center;'>CARGAR PLANILLA DE EXCEL PARA FILTRAR CAJAS</h2>", unsafe_allow_html=True)
        
            #Cargar el archivo de excel 
            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])

        
            if uploaded_file is not None:
                # Leer el archivo Excel usando Pandas
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                # Elimoinar Columnas
                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)

                # Definimos una ruta para guardar nuestros archivos
                ruta = st.text_input("Introduce la ruta de la carpeta para guardar los Archivos: Obligatorio", "C:\\")
            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])

                Nivel = st.selectbox("Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Definir la ruta de guardado
                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "2":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "3":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "4":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "5":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "6":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001", "L-PREDESP_IN", "L-PREDESP_EX", "L-ING-CJ-001", "L-INV-CJ-001", "L-SCN-CJ-001", "L-DIG-CJ-001", "L-PALLET"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV", "PREDESP_IN", "PREDESP_EX", "ING", "INV", "SCN", "DIG", "PALLET"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "PREDESP_IN":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PREDESP_IN" and Loc == "PREDESP_IN":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")
                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "PREDESP_EX":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PREDESP_EX" and Loc == "PREDESP_EX":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "ING":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-ING-CJ-001" and Loc == "ING":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "INV":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-INV-CJ-001" and Loc == "INV":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "SCN":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-SCN-CJ-001" and Loc == "SCN":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "DIG":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-DIG-CJ-001" and Loc == "DIG":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                
                if Loc == "PALLET":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PALLET" and Loc == "PALLET":
                        # Limpiar el nombre del archivo para evitar caracteres problemáticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel físico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones       
            st.markdown("<h2 style='text-align: center;'>SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar múltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Función para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Cajas.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")


        
            # Título de la aplicación      
            st.markdown("<h2 style='text-align: center;'>ELIMINAR ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Especificar la ruta de la carpeta donde están los archivos Excel
            folder_path = st.text_input("Introduce la ruta de la carpeta done Guardo Los Archivos Filtrados", "C:\\")

            # Comprobar si la ruta es válida y es una carpeta
            if folder_path and os.path.exists(folder_path) and os.path.isdir(folder_path):
                # Listar todos los archivos Excel en la carpeta
                excel_files = glob.glob(os.path.join(folder_path, "*.xlsx")) + glob.glob(os.path.join(folder_path, "*.xls"))

                # Mostrar la cantidad de archivos Excel encontrados
                st.write(f"Se encontraron {len(excel_files)} archivos Excel en la carpeta.")

                # Si hay archivos Excel, proporcionar la opción de eliminarlos
                if excel_files:
                    # Botón de confirmación para eliminar todos los archivos
                    if st.button("Eliminar todos los archivos Excel"):
                        try:
                            # Eliminar cada archivo encontrado
                            for file in excel_files:
                                os.remove(file)
                            st.success(f"Se eliminaron {len(excel_files)} archivos Excel de la carpeta.")
                        except Exception as e:
                            st.error(f"Error al eliminar archivos: {e}")
                else:
                    st.write("No se encontraron archivos Excel en la carpeta especificada.")
            else:
                st.write("Introduce una ruta válida para la carpeta.")

            # ---------------------------------------------------------------------------------------------------------

elif proyectos == "CALENDARIO ALMACEN":
    
    import io
    
    st.markdown("<h1 style='text-align: center;'>CALENDARIO ALMACEN LA PAZ</h1>", unsafe_allow_html=True)

    st.markdown("<h2 style='text-align: center;'> CARGAR PLANILLA DE EXCEL PARA REALIZAR EL CALENDARIO</h2>", unsafe_allow_html=True)

    #Cargar el archivo de excel 
    uploaded_file = st.file_uploader('SUBE TU ARCHIVO DE EXCEL', type=['xlsx','xls'])

    if uploaded_file is not None:
    # Leer el archivo Excel usando Pandas
        df = pd.read_excel(uploaded_file, engine='openpyxl')

            # Elimoinar Columnas
        Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8','ELIMINAR_9','ELIMINAR_10','ELIMINAR_11','ELIMINAR_12','ELIMINAR_13'], axis=1)
        Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
        Separar.columns = ['GALPON','LA','PASILLO','SHELF','NIVEL','COLUMNA']
        Eliminar = pd.concat([Separar, Eliminar], axis=1)
        Eliminar = Eliminar.drop(['LOCACION', 'LA'], axis=1)


        #st.dataframe(Eliminar)
        co1, co2 = st.columns(2)

        with co1:
            opciones = ["SELECCIONA UNA OPCION", "CALENDARIO GENERAL", "CALENDARIO POR PASILLO", "CALENDARIO POR SHELF"]

            seleccion = st.selectbox("Selecciona una opcion del menu: ", opciones)

        with co2:
            ruta = st.text_input("Introduce la ruta de la carpeta: ", "C:\\")

        # Función para resaltar valores mayores a 10
        def highlight_integers(values):
            if values == 0:
                return 'background-color: green'
            elif values in range(1, 12):
                return 'background-color: orange'
            elif values in range(13, 25):
                return 'background-color: red'
            return ''
        

        if seleccion == "SELECCIONA UNA OPCION":   
            pass

        elif seleccion == "CALENDARIO GENERAL":

            Calendario = Eliminar.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
            
            # Reemplazar los valores None con 0
            Calendario = Calendario.fillna(0).astype(int)

            
            #----******************************************************---
            #SACAR DATOS DEL CALENDARIO 
            # 1. TOTAL_CELDAS: filas por galpón * columnas
            total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
            total_celdas = total_celdas.rename('TOTAL_SHELS')
            # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
            nonzero_mask = (Calendario != 0)
            total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
            total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
            # 3. ESPACIOS_LIBRES: suma de ceros
            zero_mask = (Calendario == 0)
            espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
            espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
            # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
            total_cajas_para_asignar = espacios_libres * 12
            total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
            # Unir todo en un solo DataFrame
            result = pd.concat([
                total_celdas,
                total_cajas_asignadas,
                espacios_libres,
                total_cajas_para_asignar
            ], axis=1).reset_index()
                
            # Contar cendas myores a 12 
            greater_than_12 = (Calendario > 12)
            # Contar celdas mayores a 12 por NIVEL
            count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

            # Resetear índice para mejor visualización
            count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
            count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

            #---*******************************************************---

            Calendario = Calendario.style.applymap(highlight_integers)
            
            st.write(Calendario)
            
            #----*****************DESCARGAR ARCHIVO*****************************

            from xlsxwriter.utility import xl_range

            # ✅ Nombre seguro para archivo
            file_name = f"CALENDARIO_GENERAL_{pd.Timestamp.now().strftime('%Y-%m-%d_%H-%M-%S')}"
            safe_file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-") + ".xlsx"

            # ✅ Quitar estilos de Calendario si los tiene
            calendario_clean = Calendario.data if hasattr(Calendario, 'data') else Calendario

            # ✅ Crear archivo Excel en memoria
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                workbook = writer.book

                # 🧾 HOJA: CALENDARIO COMPLETO
                sheetname = 'Calendario Completo'
                calendario_clean.to_excel(writer, sheet_name=sheetname, index=True)
                worksheet = writer.sheets[sheetname]

                # ❄️ Congelar primera fila y 3 primeras columnas
                worksheet.freeze_panes(2, 4)

                # 📐 Determinar dimensiones
                n_rows, n_cols = calendario_clean.shape
                n_index_levels = calendario_clean.index.nlevels

                start_row = 1  # headers
                start_col = n_index_levels  # después de las columnas de índice

                end_row = start_row + n_rows - 1
                end_col = start_col + n_cols - 1

                cell_range = xl_range(start_row, start_col, end_row, end_col)

                # 🎨 Formatos condicionales
                formato_verde = workbook.add_format({'bg_color': '#92D050', 'border': 1})
                formato_naranja = workbook.add_format({'bg_color': '#FFC000', 'border': 1})
                formato_rojo = workbook.add_format({'bg_color': '#FF0000', 'border': 1})

                # ✅ Condiciones de color
                worksheet.conditional_format(cell_range, {
                    'type': 'cell',
                    'criteria': '==',
                    'value': 0,
                    'format': formato_verde
                })

                worksheet.conditional_format(cell_range, {
                    'type': 'cell',
                    'criteria': 'between',
                    'minimum': 1,
                    'maximum': 11,
                    'format': formato_naranja
                })

                worksheet.conditional_format(cell_range, {
                    'type': 'cell',
                    'criteria': '>=',
                    'value': 12,
                    'format': formato_rojo
                })

                # 📏 Ajuste automático de columnas
                for i, col in enumerate(calendario_clean.columns):
                    header = ' '.join(map(str, col)) if isinstance(col, tuple) else str(col)
                    max_len = max(len(header), 5)
                    worksheet.set_column(i + n_index_levels, i + n_index_levels, max_len + 2)

                # 🔄 Ajuste de ancho para columnas de índice
                for i in range(n_index_levels):
                    index_name = calendario_clean.index.names[i] or f"Index_{i}"
                    worksheet.set_column(i, i, len(index_name) + 4)

                # 🧾 HOJAS ADICIONALES
                result.to_excel(writer, index=False, sheet_name='Datos por Galpón')
                count_greater_than_12_by_nivel.to_excel(writer, index=False, sheet_name='Asignaciones en Exceso')
                Eliminar.to_excel(writer, index=False, sheet_name='Base de Datos')

            # 🧷 Finalizar el archivo
            output.seek(0)

            # 📥 Botón de descarga en Streamlit
            st.download_button(
                label="📥 Descargar Excel con Formato",
                data=output.getvalue(),
                file_name=safe_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # 💾 Guardar localmente si se da ruta
            try:
                if ruta:
                    save_path = Path(ruta) / safe_file_name
                    with open(save_path, 'wb') as f:
                        f.write(output.getvalue())
                    st.success(f"✅ Archivo guardado localmente en: {save_path.resolve()}")
            except Exception as e:
                st.warning(f"⚠️ No se pudo guardar localmente. Error: {str(e)}")

            
            # Mostrar en Streamlit

            colum1, colum2 = st.columns([12.04, 14.04])

            with colum1:
                #st.write("ASIGNACIONES QUE SOBREPASAN SU CANTIDAD LIMITE")
                st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)  
                st.table(count_greater_than_12_by_nivel)

            with colum2:
                #st.write("DATOS POR GAPON:")
                st.markdown("<h3 style='text-align: center;'>DATOS POR GAPON</h3>", unsafe_allow_html=True) 
                st.table(result)

            #---*******************************************************---

            #---*******************************************************---
            #  FILTRAR CODIGOS DE UN SHELF INDICADO
            #---*******************************************************---
            st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

            # Crear un selector para elegir una celda
            galpones = Calendario.index.get_level_values('GALPON').unique()
            pasillos = Calendario.index.get_level_values('PASILLO').unique()
            niveles = Calendario.index.get_level_values('NIVEL').unique()
            shelfs = Calendario.columns.levels[0].tolist()
            columnas = Calendario.columns.levels[1].tolist()
            
            col1, col2, col3, col4, col5 = st.columns(5)

            with col1:
                selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
            with col2:
                selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
            with col3:
                selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
            with col4:
                selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
            with col5:
                selected_columna = st.selectbox("Selecciona una Columna:", columnas)
            # Filtrar los códigos poly según la selección
            if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                filtered_data = Eliminar[
                    (Eliminar['GALPON'] == selected_galpon) &
                    (Eliminar['PASILLO'] == selected_pasillo) &
                    (Eliminar['NIVEL'] == selected_nivel) &
                    (Eliminar['SHELF'] == selected_shelf) &
                    (Eliminar['COLUMNA'] == selected_columna)
                ]

                # Asegurarse de que solo se muestren las columnas deseadas
                columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                # Renumerar los registros comensando desde 1
                filtered_data.index = filtered_data.index +1

                # Mostrar los códigos poly en una tabla
                st.write("Códigos Poly asignados a la celda seleccionada:")
                st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

            #---********************************************************---

        elif seleccion == "CALENDARIO POR PASILLO":

            file_name = st.selectbox("Guardar Como:", options = ["OPCIONES","PASILLO_A", "PASILLO_B", "PASILLO_C", "PASILLO_D", "PASILLO_E", "PASILLO_F", "PASILLO_G", "PASILLO_H", "PASILLO_I"])
                        
            Pasillo = st.selectbox("Buscar Pasillo", options = ["PASILLO", "A", "B", "C", "D", "E", "F", "G", "H", "I"])

            if Pasillo == "PASILLO":
                pass

            if Pasillo == "A":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_A" and Pasillo == "A":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.write("DATOS POR GAPON:")
                    st.table(result)

                #---*******************************************************---              

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---
    
            if Pasillo == "B":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)

                #---********************************************************---
    
                if file_name == "PASILLO_B" and Pasillo == "B":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.write("DATOS POR GAPON:")
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

            #---*******************************************************************---            

            if Pasillo == "C":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')

                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)


                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---


                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_C" and Pasillo == "C":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "D":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_D" and Pasillo == "D":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "E":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---


                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_E" and Pasillo == "E":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATO POR GALPON</h3>", unsafe_allow_html=True) 
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "F":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)

                
                if file_name == "PASILLO_F" and Pasillo == "F":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---


            if Pasillo == "G":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_G" and Pasillo == "G":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            
            if Pasillo == "H":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---


                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)

                
                if file_name == "PASILLO_H" and Pasillo == "H":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "I":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galpón * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear índice para mejor visualización
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_I" and Pasillo == "I":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardarán los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardarán los archivos


                    # Guardar el DataFrame en un archivo Excel físico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los códigos poly según la selección
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los códigos poly en una tabla
                    st.write("Códigos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---


        elif seleccion == "CALENDARIO POR SHELF":

            Pasillo = st.selectbox("Buscar Pasillo", options = ["PASILLO", "A", "B", "C", "D", "E", "F"])

            Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

            Shelf = st.selectbox("Buscar Shelf", options = ["TODOS","01", "02", "03", "04", "05", "06", "07", "08", "09"]) 

         
           # Shelf = st.text_input("Ingrese Numero de Shelf: ")

            Calendario = Calendario[(Calendario["SHELF"] == Shelf)]

            # Crear una tabla dinamica para mostrar el calendario

            Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')

            # Reemplazar los valores None con 0
            Calendario = Calendario.fillna(0).astype(int)

            #----******************************************************---
            #SACAR DATOS DEL CALENDARIO 
            # 1. TOTAL_CELDAS: filas por galpón * columnas
            total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
            total_celdas = total_celdas.rename('TOTAL_SHELS')
            # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
            nonzero_mask = (Calendario != 0)
            total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
            total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
            # 3. ESPACIOS_LIBRES: suma de ceros
            zero_mask = (Calendario == 0)
            espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
            espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
            # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
            total_cajas_para_asignar = espacios_libres * 12
            total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
            # Unir todo en un solo DataFrame
            result = pd.concat([
                total_celdas,
                total_cajas_asignadas,
                espacios_libres,
                total_cajas_para_asignar
            ], axis=1).reset_index()
                
            # Contar cendas myores a 12 
            greater_than_12 = (Calendario > 12)
            # Contar celdas mayores a 12 por NIVEL
            count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

            # Resetear índice para mejor visualización
            count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
            count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

            #---*******************************************************---

            Calendario = Calendario.style.applymap(highlight_integers)

            st.write(Calendario)

            #---********************************************************---
            # Mostrar en Streamlit
            colum1, colum2 = st.columns([12.04, 14.04])
            with colum1:
                st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                st.table(count_greater_than_12_by_nivel)
            with colum2:
                st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                st.table(result)

            #---*******************************************************---

            #---*******************************************************---
            #  FILTRAR CODIGOS DE UN SHELF INDICADO
            #---*******************************************************---
            st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

            # Crear un selector para elegir una celda
            galpones = Calendario.index.get_level_values('GALPON').unique()
            pasillos = Calendario.index.get_level_values('PASILLO').unique()
            niveles = Calendario.index.get_level_values('NIVEL').unique()
            shelfs = Calendario.columns.levels[0].tolist()
            columnas = Calendario.columns.levels[1].tolist()
                
            col1, col2, col3, col4, col5 = st.columns(5)

            with col1:
                selected_galpon = st.selectbox("Selecciona un Galpón:", galpones)
            with col2:
                selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
            with col3:
                selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
            with col4:
                selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
            with col5:
                selected_columna = st.selectbox("Selecciona una Columna:", columnas)
            # Filtrar los códigos poly según la selección
            if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                filtered_data = Eliminar[
                    (Eliminar['GALPON'] == selected_galpon) &
                    (Eliminar['PASILLO'] == selected_pasillo) &
                    (Eliminar['NIVEL'] == selected_nivel) &
                    (Eliminar['SHELF'] == selected_shelf) &
                    (Eliminar['COLUMNA'] == selected_columna)
                ]

                # Asegurarse de que solo se muestren las columnas deseadas
                columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                # Renumerar los registros comensando desde 1
                filtered_data.index = filtered_data.index +1

                # Mostrar los códigos poly en una tabla
                st.write("Códigos Poly asignados a la celda seleccionada:")
                st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

elif proyectos == "CREAR RUTA":
        import streamlit as st
        import pandas as pd
        import os
        from io import BytesIO

        # --------------------- SIDEBAR MENU ---------------------
        st.sidebar.title("Menú de Opciones")
        menu = st.sidebar.radio("Ir a:", [
            "🏷️ Filtrar Solicitudes",
            "🧩 Combinar Archivos",
            "🔎 Generar Sintaxis de Búsqueda",
            "📦 Preparar Ruta Almacén La Paz",
            "📦 Preparar Guias de Ruta",
            "🗑️ Eliminar Archivos"

        ])
        
        # --------------------- 1. FILTRAR SOLICITUDES ---------------------
        def filtrar_solicitudes():

            st.title("📊 FILTRADOR DE SOLICITUDES PARA ELABORAR RUTAS")

            co1,co2 = st.columns([15.04, 8.05])
            with co1:
                # Subida del archivo
                uploaded_file = st.file_uploader("📁 CARGAR ARCHIVO EXCEL", type=["csv","xlsx"])
            with co2:
                # Input para ruta personalizada (opcional)
                carpeta_guardado = st.text_input(
                    "📁 Ingresa la ruta para para guardar el Archivo Filtrado :",
                    placeholder="Ej: C:/Users/TuUsuario/Desktop"
                )

            # Definiciones de filtros
            TIPOS_SOLICITUD = [
                "SE1 - Envío de Materiales",
                "SR1 - Recojo de items nuevos (Cajas o Files)",
                "SR2 - Recojo de Items por devolución (Cajas o Files)"
            ]

            # Diccionario de Centros de Costo
            CENTROS_CENTRALIZADO = {
                "REGIONAL LA PAZ": {
                    "ZONA SUR": [
                        "212 - COTA COTA", 
                        "208 - SAN MIGUEL", 
                        "219 - OBRAJES"
                        ],
                    "ZONA ESTE": [
                        "204 - MIRAFLORES", 
                        "218 - VILLA ARMONIA", 
                        "211 - CRUCE VILLA COPACABANA", 
                        "223 - PAMPAHASI"],
                    "CENTRO": [
                        "221 - LA PORTADA", 
                        "205 - EL TEJAR", 
                        "216 - GRAN PODER", 
                        "202 - GARITA", "206 - ALONSO DE MENDOZA",
                        "213 - NORMALIZADORA LA PAZ", 
                        "214 - SOL AMIGO LA PAZ", 
                        "220 - TEMBLADERANI", 
                        "201 - SAN PEDRO",
                        "295 - OFICINA NACIONAL", 
                        "210 - CAMACHO", 
                        "209 - BALLIVIAN", 
                        "200 - REGIONAL LA PAZ"],
                    "NORTE": [
                        "203 - VILLA FATIMA", 
                        "224 - CHUQUIAGUILLO", 
                        "222 - PERIFERICA", 
                        "217 - VINO TINTO"]
                },
                "REGIONAL EL ALTO": {
                    "NORTE": [
                        "253 - RIO SECO", 
                        "276 - VILLA INGENIO", 
                        "272 - SAN ROQUE", 
                        "266 - MERCADO EL CARMEN RÍO SECO",
                        "275 - FRANZ TAMAYO", 
                        "251 - 16 DE JULIO", 
                        "279 - CHACALTAYA", 
                        "277 - FERROPETROL"],
                    "CENTRO": [
                        "270 - 12 DE OCTUBRE", 
                        "252 - LA CEJA", 
                        "263 - VILLA DOLORES", 
                        "258 - NORMALIZACIÓN EL ALTO",
                        "262 - SATELITE",
                        "264 - SOL AMIGO EL ALTO",
                        "265 - AGENCIA MOVIL",  
                        "250 - REGIONAL EL ALTO"],
                    "OESTE": [
                        "254 - VILLA ADELA", 
                        "261 - BOLIVIA", 
                        "274 - QUISWARAS"],
                    "SUR": [
                        "267 - SANTIAGO II", 
                        "273 - EL KENKO", 
                        "260 - SENKATA", 
                        "269 - VENTILLA"],
                    "VIACHA": [
                        "271 - INGAVI", 
                        "256 - VIACHA"]
                },
                "REGIONAL ORURO": {
                    "ORURO": [
                        "401 - CENTRAL", 
                        "407 - NORMALIZACION ORURO", 
                        "408 - PUNTO AMIGO ORURO", 
                        "409 - VIRGEN DEL SOCAVÓN",
                        "410 - TAGARETE", 
                        "411 - TACNA"]
                },
                "REGIONAL SUCRE": {
                    "SUCRE": [
                        "100 - REGIONAL CHUQUISACA", 
                        "101 - MERCADO CAMPESINO SUCRE", 
                        "102 - NORMALIZACION SUCRE", 
                        "103 - ESPAÑA",
                        "104 - SOL AMIGO SUCRE", 
                        "105 - GERMAN MENDOZA", 
                        "106 - CHARCAS", 
                        "107 - ZUDAÑEZ",
                        "108 - LAS AMERICAS", 
                        "109 - LAJASTAMBO"]
                },
                "REGIONAL TARIJA": {
                    "": [
                        "601 - MERCADO CAMPESINO TARIJA",
                        "603 - SUR",
                        "604 - SOL AMIGO TARIJA",
                        "605 - 15 DE ABRIL",
                        "606 - NORMALIZACIÓN TARIJA",
                        "607 - TABLADITA",
                        "608 - YACUIBA",
                        "609 - PALMARCITO",
                        "610 - MERCADO CAMPESINO YACUIBA"
                    ]
                },

                "BANCO NACIONAL DE BOLIVIA": {
                    "BNB - LA PAZ": [
                        "130 - BANCO NACIONAL DE BOLIVIA - LA PAZ"],
                    "BNB - COCHABAMBA":[
                        "132 - BANCO NACIONAL DE BOLIVIA - COCHABAMBA"],
                    "BNB - TARIJA":[
                        "136 - BANCO NACIONAL DE BOLIVIA - TARIJA"],
                    "BNB - ORURO":[
                        "134 - BANCO NACIONAL DE BOLIVIA - ORURO"],
                    "BNB - SUCRE":[
                        "133 - BANCO NACIONAL DE BOLIVIA - SUCRE"],
                    "BNB - SANTA CRUZ":[
                        "131 - BANCO NACIONAL DE BOLIVIA - SANTA CRUZ"],
                    "BNB - BENI":[
                        "137 - BANCO NACIONAL DE BOLIVIA - BENI"],
                    "BNB - POTOSI":[
                        "135 - BANCO NACIONAL DE BOLIVIA - POTOSI"],
                    "BNB - PANDO":[
                        "138 - BANCO NACIONAL DE BOLIVIA - PANDO"],
                    "BNB - POLYSISTEMAS":[
                        "C008 - POLYSISTEMAS"]

                },

                "BANCO FIE": {
                    "BFIE LA PAZ": [
                        "124-1 - BFIE LA PAZ"],
                    "BFIE SANTA CRUZ": [
                        "124-2 - BFIE SANTA CRUZ"],
                    "BFIE INTERNO": [   
                        "C008 - POLYSISTEMAS"]
                        
                },

                "JTI BOLIVIA": {
                    "JTI EXTERNO": [
                        "367 - JTI BOLIVIA"],
                    "JTI INTERNO": [   
                        "C008 - POLYSISTEMAS"]
                },

                "REGIONAL COCHABAMBA": {
                    "": [
                        "301 - ESTEBAN ARCE",
                        "302 - SAN MARTIN",
                        "303 - HUAYRA KHASA",
                        "305 - CRUCE TAQUIÑA",
                        "306 - QUILLACOLLO",
                        "307 - COLCA PIRHUA",
                        "309 - MUYURINA",
                        "310 - NORMALIZACIÓN COCHABAMBA",
                        "311 - RECAUDADORA JORDAN",
                        "312 - SACABA",
                        "313 - VILLA GALINDO",
                        "314 - PUNATA",
                        "316 - AYACUCHO",
                        "317 - SOL AMIGO COCHABAMBA",
                        "318 - PANAMERICANA",
                        "320 - CLIZA",
                        "321 - VINTO",
                        "322 - REPÚBLICA",
                        "323 - TIQUIPAYA",
                        "324 - QUINTANILLA",
                        "325 - JORDAN",
                        "326 - PLAZA BOLIVAR",
                        "327 - PETROLERA",
                        "328 - LA CHIMBA",
                        "329 - AMERICA",
                        "331 - EL AVION",
                        "332 - VILLA PAGADOR",
                        "333 - PACATA"
                    ]
                },
                "REGIONAL BENI": {
                    "": [
                        "801 - TRINIDAD",
                        "802 - RIBERALTA",
                        "803 - GUAYARAMERIN"
                    ]
                },
                "REGIONAL COBIJA": {
                    "": [
                        "901 - COBIJA",
                        "902 - TAJIBOS",
                        "903 - NORMALIZACIÓN"
                    ]
                },
                "REGIONAL POTOSI": {
                    "": [
                        "501 - MERCADO UYUNI",
                        "502 - SOL AMIGO POTOSI",
                        "503 - BOULEVARD"
                    ]
                },
                "REGIONAL SANTA CRUZ": {
                    "": [
                        "701 - CASCO VIEJO",
                        "702 - EL PARI",
                        "703 - MUTUALISTA",
                        "704 - 1RO. DE MAYO",
                        "705 - MONTERO",
                        "706 - EL TORNO",
                        "709 - PIRAI",
                        "711 - PLAN 3000",
                        "715 - LA GUARDIA",
                        "716 - ALTO SAN PEDRO",
                        "718 - NORTE",
                        "719 - SOL AMIGO SANTA CRUZ",
                        "721 - ARROYO CONCEPCIÓN",
                        "723 - PAMPA DE LA ISLA",
                        "724 - COLORADA",
                        "725 - Minero",
                        "726 - SAN JULIAN",
                        "727 - SAN JOSE",
                        "728 - 2 DE AGOSTO",
                        "730 - LOS LOTES",
                        "731 - GERMAN MORENO",
                        "732 - YAPACANI",
                        "734 - SATELITE NORTE",
                        "735 - NORTE I",
                        "736 - EL QUIOR",
                        "737 - VIRGEN DE LUJAN",
                        "738 - EL BAJIO",
                        "739 - CRISTO REDENTOR"
                    ]
                }
            }

            if uploaded_file:
                df = pd.read_excel(uploaded_file)
                
                col1, col2 = st.columns(2)

                with col1:
                    st.subheader("FILTRAR TIPO DE SOLICITUD")
                    tipos_seleccionados = st.multiselect(
                        "Selecciona uno o varios tipos de solicitud:",
                        options=TIPOS_SOLICITUD,
                        default=TIPOS_SOLICITUD
                    )
                with col2:
                    st.subheader("SELECCIONAR REGION Y SUBZONA")

                    region = st.selectbox("Selecciona una región", options=list(CENTROS_CENTRALIZADO.keys()))
                    subzonas = list(CENTROS_CENTRALIZADO[region].keys())
                    subzona = st.selectbox("Selecciona una subzona", options=subzonas)

                    centros_disponibles = CENTROS_CENTRALIZADO[region][subzona]
                    centros_seleccionados = st.multiselect("Selecciona Centros de Costo", centros_disponibles, default=centros_disponibles)

                # Filtrado de datos
                df_filtrado = df[
                    (df["Tipo de Solicitud"].isin(tipos_seleccionados)) &
                    (df["Centro de Costo"].isin(centros_seleccionados))
                ]

                # Eliminar columnas no deseadas
                columnas_a_eliminar = ["Autorizado", "Locacion", "Centro de Costo Polysistemas", "Fecha de Impresión"]
                df_filtrado = df_filtrado.drop(columns=columnas_a_eliminar, errors='ignore')

                st.success(f"🔍 Filtrado completo: {len(df_filtrado)} registros encontrados.")

                colum1, colum2 = st.columns([15.50, 6.10])

                with colum1:
                    # Mostrar tabla filtrada
                    st.dataframe(df_filtrado, use_container_width=True)

                with colum2:
                    # Mostrar resumen basado en TipoFile y suma de Cantidad como tabla
                    tipos_resumen = ["CAJA", "Caja", "Cintillos", "FILE"]
                    st.subheader("📦 RESUMEN DE CANTIDADES")

                    if "TipoFile" in df_filtrado.columns and "Cantidad" in df_filtrado.columns:
                        resumen_data = []

                        for tipo in tipos_resumen:
                            total = df_filtrado[df_filtrado["TipoFile"] == tipo]["Cantidad"].sum()
                            resumen_data.append({"TipoFile": tipo, "Total Cantidad": int(total)})

                        resumen_df = pd.DataFrame(resumen_data)
                        st.table(resumen_df)
                    else:
                        st.warning("⚠️ Las columnas 'TipoFile' o 'Cantidad' no se encuentran en el archivo.")

                nombre_archivo = "resultado_filtrado.xlsx"  # nombre fijo

                if st.button("📥 Descargar Excel filtrado / Guardar archivo"):
                    if carpeta_guardado.strip():
                        ruta = os.path.join(carpeta_guardado.strip(), nombre_archivo)
                        try:
                            if not os.path.exists(carpeta_guardado):
                                os.makedirs(carpeta_guardado)  # crea carpeta si no existe
                            df_filtrado.to_excel(ruta, index=False)
                            st.success(f"✅ Archivo guardado correctamente en:\n`{ruta}`")
                        except Exception as e:
                            st.error(f"❌ Error al guardar archivo: {e}")
                    else:
                        # Si no ingresa carpeta, descarga por navegador
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_filtrado.to_excel(writer, index=False, sheet_name='Filtrado')
                        output.seek(0)
                        st.download_button(
                            label="📥 Descargar Excel filtrado",
                            data=output,
                            file_name=nombre_archivo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

        # --------------------- 2. COMBINAR ARCHIVOS ---------------------
        def combinar_archivos():
            st.title("📚 COMBINAR ARCHIVOS FILTRADOS")
              

            co1,co2 = st.columns([15.04, 8.05])
            with co1:
                # Cargar múltiples archivos
                uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)
                
            with co2:
                # Input para ruta personalizada (opcional)
                carpeta_guardado = st.text_input(
                    "📁 Ingresa la ruta para para guardar el Archivo Filtrado :",
                    placeholder="Ej: C:/Users/TuUsuario/Desktop"
                )

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)

                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Función para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output
                
                nombre_archivo2 = "Archivo_Combinado.xlsx"  # nombre fijo

                if st.button("📥 Descargar Excel Combinado / Guardar archivo"):
                    if carpeta_guardado.strip():
                        ruta = os.path.join(carpeta_guardado.strip(), nombre_archivo2)
                        try:
                            if not os.path.exists(carpeta_guardado):
                                os.makedirs(carpeta_guardado)  # crea carpeta si no existe
                            combined_df.to_excel(ruta, index=False)
                            st.success(f"✅ Archivo guardado correctamente en:\n`{ruta}`")
                        except Exception as e:
                            st.error(f"❌ Error al guardar archivo: {e}")
                
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")


        # --------------------- 3. GENERADOR DE SINTAXIS ---------------------
        def generar_sintaxis():

            st.title("🔎 GENERADOR DE SINTAXIS DE BÚSQUEDA")

            col1, col2 = st.columns((2))
            with col1:
                # Cargar archivo Excel
                st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCode'</h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):

                    if uploaded_file is not None:
                        try:
                            # Leer el archivo Excel y obtener solo la columna 'SolicitudCode'
                            df = pd.read_excel(uploaded_file)

                            # Verificar si la columna existe
                            if 'SolicitudCode' not in df.columns:
                                st.error("La columna 'SolicitudCode' no se encuentra en el archivo.")
                            else:
                                # Eliminar las demás columnas y eliminar nulos
                                solicitud_codes = df['SolicitudCode'].dropna().astype(str).tolist()

                                # Crear la sintaxis
                                solicitud_sintaxis = " | ".join(
                                    [f'\n{{[Solicitud]:[Nro Solicitud]="{code.strip()}"}}' for code in solicitud_codes if code.strip()]
                                )

                                sintaxis_final = f"{solicitud_sintaxis}"
                        except Exception as e:
                            st.error(f"Error al leer el archivo: {e}")

            with col2:
                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)

       
        # --------------------- 4. PREPARAR RUTA ALMACÉN ---------------------
        def preparar_ruta():

            import streamlit as st
            import pandas as pd
            from io import BytesIO
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            st.title("📦 PREPARAR RUTA ALMACÉN LA PAZ")

            col1, col2 = st.columns(2)
            file1 = col1.file_uploader("📂 Cargar Archivo 1", type="xlsx")
            file2 = col2.file_uploader("📂 Cargar Archivo 2", type="xlsx")

            file_name_input = st.text_input("📝 Nombre del archivo de descarga (sin .xlsx)", value="archivo_Ruta_Final")
            sheet_name_input = st.text_input("📄 Nombre de la hoja en el archivo Excel", value="Ruta")

            recojo_title = st.text_input("✏️ Título para la sección de Recojo", value="RECOJO DE CAJAS B_SOL")
            devolucion_title = st.text_input("✏️ Título para la sección de Devolución", value="DEVOLUCIÓN DE ITEMS B_SOL")

            if file1 and file2:
                colum1, colum2 = st.columns(2)

                with colum1:
                    df1 = pd.read_excel(file1)
                    df1.drop(['Estado de Solicitud', 'Servicio', 'Turno', 'Tipo de Recojo', 'Fecha de Registro'], axis=1, inplace=True, errors='ignore')
                    st.write("Contenido del primer archivo:")
                    st.dataframe(df1.head())

                with colum2:
                    df2 = pd.read_excel(file2)
                    st.write("Contenido del segundo archivo:")
                    st.dataframe(df2.head())

                # Aseguramos que claves para merge sean str sin espacios extra
                df1['SolicitudCode'] = df1['SolicitudCode'].astype(str).str.strip()
                df2['Nro Solicitud'] = df2['Nro Solicitud'].astype(str).str.strip()

                if 'SolicitudCode' not in df1.columns or 'Nro Solicitud' not in df2.columns:
                    st.error("❌ Faltan columnas clave en los archivos ('SolicitudCode' o 'Nro Solicitud').")
                else:
                    merged_df = pd.merge(
                        df1,
                        df2[['Nro Solicitud', 'Solicitante', 'Usuario', 'Items Oneil']],
                        left_on='SolicitudCode',
                        right_on='Nro Solicitud',
                        how='left'
                    )

                    # Mostramos las primeras 5 filas del archivo combinado
                    st.subheader(" Vista previa del archivo combinado")
                    st.dataframe(merged_df.head())

                    if 'Centro de Costo' in merged_df.columns:
                        merged_df = merged_df.sort_values(by='Centro de Costo')

                    df_recojo = merged_df[merged_df['TipoFile'] != 'FILE'].copy()
                    df_devolucion = merged_df[merged_df['TipoFile'] == 'FILE'].copy()

                    cols_to_drop = ['Usuario', 'Nro Solicitud']
                    df_recojo.drop(columns=cols_to_drop, inplace=True, errors='ignore')
                    df_devolucion.drop(columns=cols_to_drop, inplace=True, errors='ignore')

                    st.subheader("📋 Ruta del Día - Recojo")
                    st.dataframe(df_recojo)
                    st.subheader("📋 Ruta del Día - Devolución")
                    st.dataframe(df_devolucion)

                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        sheet_name = sheet_name_input.strip() or "Ruta"
                        workbook = writer.book
                        worksheet = workbook.create_sheet(sheet_name)
                        if 'Sheet' in workbook.sheetnames:
                            std = workbook['Sheet']
                            workbook.remove(std)

                        color_lila = "C27BA0"
                        color_rosa = "F4B1BA"
                        bold_font = Font(bold=True)
                        thin_border = Border(
                            left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000")
                        )

                        def apply_borders(ws, start_row, df):
                            nrows, ncols = df.shape
                            for row in range(start_row, start_row + nrows + 1):  # +1 para incluir header
                                for col in range(1, ncols + 1):
                                    ws.cell(row=row, column=col).border = thin_border

                        current_row = 1

                        # --- Recojo ---
                        last_col_recojo = get_column_letter(len(df_recojo.columns) if len(df_recojo.columns) > 0 else 1)
                        worksheet.merge_cells(f'A{current_row}:{last_col_recojo}{current_row}')
                        cell = worksheet[f'A{current_row}']
                        cell.value = recojo_title
                        cell.font = Font(bold=True, size=14)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill("solid", fgColor=color_lila)

                        if not df_recojo.empty:
                            startrow = current_row + 2
                            df_recojo.to_excel(writer, index=False, startrow=startrow, sheet_name=sheet_name)
                            header_row = startrow + 1

                            for col in range(1, len(df_recojo.columns) + 1):
                                cell = worksheet.cell(row=header_row, column=col)
                                cell.fill = PatternFill("solid", fgColor=color_lila)
                                cell.font = bold_font

                            for i, col_name in enumerate(df_recojo.columns, 1):
                                width = max(df_recojo[col_name].astype(str).map(len).max(), len(col_name)) + 2
                                worksheet.column_dimensions[get_column_letter(i)].width = width

                            apply_borders(worksheet, header_row, df_recojo)
                            current_row = header_row + len(df_recojo) + 3
                        else:
                            startrow = current_row + 2
                            worksheet.cell(row=startrow, column=1).value = "No hay datos para recojo"
                            current_row = startrow + 3

                        # --- Devolución ---
                        last_col_dev = get_column_letter(len(df_devolucion.columns) if len(df_devolucion.columns) > 0 else 1)
                        worksheet.merge_cells(f'A{current_row}:{last_col_dev}{current_row}')
                        cell = worksheet[f'A{current_row}']
                        cell.value = devolucion_title
                        cell.font = Font(bold=True, size=12)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill("solid", fgColor=color_rosa)

                        if not df_devolucion.empty:
                            startrow = current_row + 2
                            df_devolucion.to_excel(writer, index=False, startrow=startrow, sheet_name=sheet_name)
                            header_row = startrow + 1

                            for col in range(1, len(df_devolucion.columns) + 1):
                                cell = worksheet.cell(row=header_row, column=col)
                                cell.fill = PatternFill("solid", fgColor=color_rosa)
                                cell.font = bold_font

                            for i, col_name in enumerate(df_devolucion.columns, 1):
                                width = max(df_devolucion[col_name].astype(str).map(len).max(), len(col_name)) + 2
                                worksheet.column_dimensions[get_column_letter(i)].width = width

                            apply_borders(worksheet, header_row, df_devolucion)
                        else:
                            startrow = current_row + 2
                            worksheet.cell(row=startrow, column=1).value = "No hay datos para devolución"
                            current_row = startrow + 3

                    output.seek(0)
                    st.download_button(
                        label="📥 Descargar Excel Personalizado",
                        data=output,
                        file_name=(file_name_input.strip() or "archivo_Ruta_Final") + ".xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )


        def preparar_guia_ruta():

            import streamlit as st
            import pandas as pd
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from datetime import datetime

            st.title("📦 Reporte Detallado por Solicitante")

            # Cargar archivos
            col1, col2, col3 = st.columns(3)
            with col1:
                uploaded_file = st.file_uploader("Cargar archivo combinado con Items Oneil", type="xlsx")
            with col2:
                logo_file = st.file_uploader("📷 Cargar primer logo", type=["png", "jpg", "jpeg"])
            with col3:
                logo_file2 = st.file_uploader("📷 Cargar segundo logo", type=["png", "jpg", "jpeg"], key="logo2")

            if uploaded_file:
                df = pd.read_excel(uploaded_file)
                df.columns = df.columns.str.strip().str.replace(" ", "_").str.replace("-", "_")

                required_columns = ['Solicitante', 'Centro_de_Costo', 'WorkOrderCode', 'Cantidad', 'Items_Oneil', 'TipoFile']
                if not all(col in df.columns for col in required_columns):
                    st.error(f"❌ Faltan columnas requeridas: {', '.join(required_columns)}")
                    st.stop()

                colum1, colum2, colum3 = st.columns(3)
                with colum1:
                    # Configuración de filtros y datos
                    tipos_disponibles = df['TipoFile'].dropna().unique()
                    tipos_seleccionados = st.multiselect("Selecciona uno o más TipoFile a procesar:", tipos_disponibles)
                    if not tipos_seleccionados:
                        st.warning("⚠️ Por favor, selecciona al menos un TipoFile para continuar.")
                        st.stop()
                with colum2:
                    df = df[df['TipoFile'].isin(tipos_seleccionados)]

                    encargado = st.selectbox("👤 Encargado:", ["", "JAIME QUISPE", "CARLOS ORTIZ", "MARCO HUAYLLUCO", "ALFREDO RIVEROS"])
                    if not encargado:
                        st.warning("⚠️ Por favor, selecciona el encargado.")
                        st.stop()
                with colum3:
                    guia_seleccionada = st.selectbox("Tipo de Guía:", ["GUÍA DE RECEPCIÓN DE CAJAS", "GUÍA DE RECOJO DE ITEMS", "GUÍA DE ENTREGA DE MATERIALES"])
                    regional_seleccionada = st.selectbox("Regional:", ["REGIONAL LA PAZ", "REGIONAL EL ALTO", "REGIONAL SUCRE", "REGIONAL ORURO", "REGIONAL POTOSI"])
                    fecha_actual = datetime.now().strftime("%d/%m/%Y")

                st.info(f"""
                    **👤 Encargado:** {encargado}  
                    **📅 Fecha del Reporte:** {fecha_actual}  
                    **📋 Tipo de Guía:** {guia_seleccionada}  
                    **📍 Regional:** {regional_seleccionada}
                    """)

                st.subheader("🔍 Vista previa de datos")
                st.write(df.head(2))

                # Preparar workbook
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Reporte Detallado"
                ws.sheet_view.showGridLines = False

                # Estilos
                bold = Font(bold=True)
                verde = PatternFill("solid", fgColor="A9D18E")
                morado = PatternFill("solid", fgColor="C27BA0")
                gris = PatternFill("solid", fgColor="D9D9D9")
                center = Alignment(horizontal="center", vertical="center")
                wrap_text = Alignment(wrap_text=True, vertical="top")
                thin = Side(style="thin")
                thick = Side(style="thick")

                # Logos
                if logo_file:
                    image_stream = BytesIO(logo_file.read())
                    logo = OpenpyxlImage(image_stream)
                    logo.width, logo.height = 190, 80
                    ws.add_image(logo, "A1")
                if logo_file2:
                    image_stream2 = BytesIO(logo_file2.read())
                    logo2 = OpenpyxlImage(image_stream2)
                    logo2.width, logo2.height = 190, 80
                    ws.add_image(logo2, "G1")

                # ENCABEZADO ENMARCADO
                encabezados = [
                    (1, guia_seleccionada),
                    (2, regional_seleccionada),
                    (3, "RESPONSABLE POLYSISTEMAS BOLIVIA S.R.L.")
                ]
                for row, texto in encabezados:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                    ws.cell(row=row, column=1, value=texto).font = bold
                    ws.cell(row=row, column=1).alignment = center
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).border = Border(
                            top=thick if row == 1 else thin,
                            bottom=thick if row == 3 else thin,
                            left=thick if col == 1 else thin,
                            right=thick if col == 7 else thin
                        )

                # Fila con ENCARGADO y FECHA
                ws.merge_cells("A5:B5")
                ws.merge_cells("C5:D5")
                ws.merge_cells("E5:F5")
                ws["A5"].value = "ENCARGADO:"
                ws["C5"].value = encargado
                ws["E5"].value = "FECHA:"
                ws["G5"].value = fecha_actual
                for col in [1, 3, 5, 7]:
                    ws.cell(row=5, column=col).font = bold
                    ws.cell(row=5, column=col).alignment = center
                for col in range(1, 8):
                    ws.cell(row=5, column=col).border = Border(
                        top=thin,
                        bottom=thick,
                        left=thick if col == 1 else thin,
                        right=thick if col == 7 else thin
                    )

                # Total cajas
                ws.merge_cells("A6:D6")
                ws["A6"] = "CANTIDAD TOTAL"
                ws["A6"].font = bold
                ws["A6"].alignment = center
                ws["E6"] = df['Cantidad'].sum()
                ws["E6"].font = bold
                ws["E6"].alignment = center
                for col in range(1, 6):
                    ws.cell(row=6, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

                current_row = 7
                df = df.dropna(subset=['Solicitante'])

                # Agrupar por Solicitante
                for solicitante, grupo_solicitante in df.groupby('Solicitante'):
                    for centro, grupo_agencia in grupo_solicitante.groupby('Centro_de_Costo'):
                        start_row = current_row
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
                        ws.cell(row=current_row, column=1, value=f"SOLICITANTE: {solicitante}").font = bold
                        ws.cell(row=current_row, column=5, value=f"AGENCIA: {centro}").font = bold
                        for col in range(1, 8):
                            ws.cell(row=current_row, column=col).fill = verde
                            ws.cell(row=current_row, column=col).alignment = center
                            ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        current_row += 1

                        headers = ["NRO", "NRO. WO", "CANTIDAD", "CONTAINER CODE", "OBSERVACIONES"]
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=current_row, column=col_num, value=header).font = bold
                            ws.cell(row=current_row, column=col_num).fill = morado
                            ws.cell(row=current_row, column=col_num).alignment = center
                            ws.cell(row=current_row, column=col_num).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        current_row += 1

                        total_cantidad = 0
                        for i, row in enumerate(grupo_agencia.itertuples(index=False), 1):
                            ws.cell(row=current_row, column=1, value=i).alignment = center
                            ws.cell(row=current_row, column=2, value=row.WorkOrderCode).alignment = center
                            ws.cell(row=current_row, column=3, value=row.Cantidad).alignment = center
                            ws.cell(row=current_row, column=4, value=str(row.Items_Oneil).replace(";", ",")).alignment = wrap_text
                            ws.cell(row=current_row, column=5, value="").alignment = wrap_text
                            for col in range(1, 6):
                                ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                            total_cantidad += row.Cantidad
                            current_row += 1

                        # Totales y Firmas
                        ws.cell(row=current_row, column=2, value="TOTAL").font = bold
                        ws.cell(row=current_row, column=3, value=total_cantidad).font = bold
                        for col in range(1, 6):
                            ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

                        ws.cell(row=current_row, column=6, value="ENTREGUE CONFORME").font = bold
                        ws.cell(row=current_row, column=6).fill = gris
                        ws.cell(row=current_row, column=6).alignment = center
                        ws.cell(row=current_row, column=6).border = Border(left=thin, right=thin, top=thin, bottom=thin)

                        ws.cell(row=current_row, column=7, value="RECIBÍ CONFORME").font = bold
                        ws.cell(row=current_row, column=7).fill = gris
                        ws.cell(row=current_row, column=7).alignment = center
                        ws.cell(row=current_row, column=7).border = Border(left=thin, right=thin, top=thin, bottom=thin)

                        # Enmarcar todo el bloque con borde grueso
                        end_row = current_row
                        for r in range(start_row, end_row + 1):
                            for c in range(1, 8):
                                cell = ws.cell(row=r, column=c)
                                cell.border = Border(
                                    top=thick if r == start_row else thin,
                                    bottom=thick if r == end_row else thin,
                                    left=thick if c == 1 else thin,
                                    right=thick if c == 7 else thin
                                )
                        current_row += 2

                # Ajuste de columnas
                for i, width in enumerate([8, 15, 10, 50, 30, 25, 25], 1):
                    ws.column_dimensions[get_column_letter(i)].width = width

                wb.save(output)
                output.seek(0)

                nombre_archivo = st.text_input("📝 Ingresa el nombre del archivo:", "Reporte_Solicitantes")
                if not nombre_archivo.strip():
                    st.warning("⚠️ Por favor, ingresa un nombre válido para el archivo.")
                else:
                    st.success("✅ Reporte generado con encabezado completo, firmas y totales.")
                    st.download_button(
                    label="📥 Descargar Reporte Final",
                    data=output,
                    file_name=f"{nombre_archivo.strip()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )



          
        # --------------------- 5. PREPARAR RUTA ALMACÉN ---------------------
        def eliminar_archivos():
            import streamlit as st
            import os
            import glob
            import pandas as pd

            #st.set_page_config(page_title="Eliminar Excel con vista previa", layout="centered")
            st.title("🗑️ Eliminar archivos Excel con vista previa desplegable")

            folder_input = st.text_input("Introduce la ruta de la carpeta con archivos Excel:", "C:\\Users\\juan.ramos\\Desktop\\FILTRADOR_ALMACEN\RUTAS_LA_PAZ")
            if st.button("🔄 Cargar archivos"):
                st.session_state.folder = folder_input
                st.session_state.selected_files = []

            # Inicializar estados
            if 'folder' not in st.session_state:
                st.session_state.folder = ""
            if 'selected_files' not in st.session_state:
                st.session_state.selected_files = []

            folder = st.session_state.folder

            def toggle_selection(f_name):
                sel = st.session_state.selected_files
                if f_name in sel:
                    sel.remove(f_name)
                else:
                    sel.append(f_name)

            if folder:
                if os.path.isdir(folder):
                    files = glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls"))
                    if files:
                        st.subheader("📂 Selecciona archivos:")
                        for path in files:
                            name = os.path.basename(path)
                            selected = name in st.session_state.selected_files
                            col1, col2 = st.columns([4,1])
                            with col1:
                                st.write(name)
                            with col2:
                                if st.button("✅" if selected else "☐", key=name):
                                    toggle_selection(name)
                                    st.rerun()

                        if st.session_state.selected_files:
                            st.subheader("👁️ Previsualizaciones desplegables")
                            for name in st.session_state.selected_files:
                                file_path = os.path.join(folder, name)
                                with st.expander(f"Ver preview: {name}", expanded=False):
                                    try:
                                        df = pd.read_excel(file_path, nrows=10)
                                        st.dataframe(df)
                                    except Exception as e:
                                        st.error(f"Error leyendo {name}: {e}")

                            if st.checkbox("⚠️ Confirmo que deseo eliminar los archivos seleccionados"):
                                if st.button("🗑️ Eliminar archivos"):
                                    cnt = 0
                                    for name in list(st.session_state.selected_files):
                                        p = os.path.join(folder, name)
                                        if os.path.exists(p):
                                            os.remove(p)
                                            cnt += 1
                                    st.success(f"Se eliminaron {cnt} archivos correctamente.")
                                    st.session_state.selected_files = []
                                    st.session_state.folder = ""
                                    st.rerun()
                        else:
                            st.info("Selecciona archivos pulsando en los botones de la derecha.")
                    else:
                        st.warning("No se encontraron archivos Excel en la carpeta.")
                else:
                    st.error("La ruta no es válida o no corresponde a una carpeta.")
            else:
                st.info("Introduce la ruta y pulsa 'Cargar archivos' para empezar.")

                                
        # --------------------- EJECUTAR SEGÚN MENÚ ---------------------
        if menu == "🏷️ Filtrar Solicitudes":
            filtrar_solicitudes()
        elif menu == "🧩 Combinar Archivos":
            combinar_archivos()
        elif menu == "🔎 Generar Sintaxis de Búsqueda":
            generar_sintaxis()
        elif menu == "📦 Preparar Ruta Almacén La Paz":
            preparar_ruta()
        elif menu == "📦 Preparar Guias de Ruta":
            preparar_guia_ruta()
        elif menu == "🗑️ Eliminar Archivos":
            eliminar_archivos()

elif proyectos == "SINTAXIS LASERFICHE-ONEIL":
    import streamlit as st
    from datetime import datetime
    
    # Título de la aplicación
    st.markdown("<h1 style='text-align: center;'>SINTAXIS PARA LASER FICHE - ONEIL </h1>", unsafe_allow_html=True)

    # Crear una lista de opciones para el menu 
    #opciones = ["SELECCIONA UNA OPCION", "OTPM - MATERIALES", "OTEX - EXTRACIONES", "OTEV - ENVIO", "OTRE - RECOJO", "OTRE - POR_WORKORDERS", "OTRE - POR_SOLICITUDES", "OTRE - SOLICITUDES", "COD - BASE_DE_DATOS", "CONTENIDO_PALLETS"]

    # Crear el menu desplegable con st.selectbox()
    #seleccion = st.selectbox("Selecciona una opcion del menu: ", opciones)
    
    # --------------------- SIDEBAR MENU ---------------------
    st.sidebar.title("Menú de Opciones")
    menu = st.sidebar.radio("Ir a:", [
        "🧱 OTPM - MATERIALES",
        "🗃️ OTEX - EXTRACIONES",
        "📤 OTEV - ENVIO",
        "📥 OTRE - RECOJO",
        "✉️ OTRE - POR_WORKORDERS",
        "🧾 OTRE - POR_SOLICITUDES",
        "📝 OTRE - SOLICITUDES",
        "🧠 COD - BASE_DE_DATO",
        "🎨 CONTENIDO_PALLETS",
        "🔖 CODIGOS DE 6 DIGITOS"
        
    ])
    def OTPM_MATERIALES():
        
        opciones2 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

        # Crear el menu desplegable con st.selectbox()
        seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

        if seleccion2 == "PENDIENTE":
            #import streamlit as st
            from datetime import datetime
            
            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\03. Finalizada\\01. Pendiente"}'

                # Sección para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

            if __name__ == "__main__":
                main()
    
        elif seleccion2 == "EN PROCESO":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\02. En Proceso"}'

                # Sección para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

            if __name__ == "__main__":
                main()

        elif seleccion2 == "FINALIZADA":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\03. Finalizada"}'

                # Sección para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

            if __name__ == "__main__":
                main()
        
        elif seleccion2 == "ANULADA":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\04. Anulada"}'

                # Sección para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

            if __name__ == "__main__":
                main()

    def OTEX_EXTRACIONES():

        # Crear una lista de opciones para el menu 
        opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

        # Crear el menu desplegable con st.selectbox()
        seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

        if seleccion3 == "PENDIENTE":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\01. Pendiente"}'

                # Sección para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

            if __name__ == "__main__":
                main()


        elif seleccion3 == "EN PROCESO":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\02. En proceso"}'

                # Sección para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

            if __name__ == "__main__":
                main()

        elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\03. Finalizada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

        elif seleccion3 == "ANULADA":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\04. Anulada"}'

                # Sección para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

            if __name__ == "__main__":
                main()

    def OTEV_ENVIO():

        # Crear una lista de opciones para el menu 
        opciones2 = ["SELECCIONA UNA OPCION", "INTERNO", "EXTERNO"]

        # Crear el menu desplegable con st.selectbox()
        seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

        if seleccion2 == "INTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\01. Pendiente"}'

                    # Sección para seleccionar una fecha usando un calendariO
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()


            elif seleccion2 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\02. En proceso"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion2 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\03. Finalizada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                        main()

            elif seleccion2 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\04. Anulada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

        if seleccion2 == "EXTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\01. Pendiente"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()


            elif seleccion3 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\02. En proceso"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\03. Finalizada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\04. Anulada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

    def OTRE_RECOJO():

        # Crear una lista de opciones para el menu 
        opciones2 = ["SELECCIONA UNA OPCION", "INTERNO", "EXTERNO"]

        # Crear el menu desplegable con st.selectbox()
        seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

        if seleccion2 == "INTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\01. Pendiente"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.header("Seleccione una fecha")
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()


            elif seleccion3 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\02. En proceso"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\03. Finalizada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\04. Anulada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

        if seleccion2 == "EXTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\01. Pendiente"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()


            elif seleccion3 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\02. En proceso"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\03. Finalizada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\04. Anulada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

    def OTRE_POR_WORKORDERS():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Nros de WOs </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        wo_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                        # Crear la parte de la sintaxis para los Nros de WO
                        wo_sintaxis = " | ".join([f'\n{{[Orden de Trabajo]:[Nro de WO]="{wo.strip()}"}}' for wo in wo_numbers if wo.strip()])

                        # Construir la sintaxis final
                        sintaxis_final = f"{wo_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
        if __name__ == "__main__":
            main()

    def OTRE_POR_SOLICITUDES():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Nros de Solicitudes </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        wo_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                        # Crear la parte de la sintaxis para los Nros de WO
                        wo_sintaxis = " | ".join([f'\n{{[Orden de Trabajo]:[Nro Solicitud]="{wo.strip()}"}}' for wo in wo_numbers if wo.strip()])

                        # Construir la sintaxis final
                        sintaxis_final = f"{wo_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
        if __name__ == "__main__":
            main()
 
    def OTRE_SOLICITUDES():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Nros de SOLICITUDES </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        solicitud_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                        # Crear la parte de la sintaxis para los Nros de WO
                        solicitud_sintaxis = " | ".join([f'\n{{[Solicitud]:[Nro Solicitud]="{solicitud.strip()}"}}' for solicitud in solicitud_numbers if solicitud.strip()])

                        # Construir la sintaxis final
                        sintaxis_final = f"{solicitud_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
        if __name__ == "__main__":
            main()

    def COD_BASE_DE_DATO():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Codigos para descargar BD </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Codigos Poly </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        solicitud_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                        # Crear la parte de la sintaxis para los Nros de WO
                        solicitud_sintaxis = " OR ".join([f'\n Filefolder^ContainerCode = "{solicitud.strip()}"' for solicitud in solicitud_numbers if solicitud.strip()])
                        # Construir la sintaxis final
                        sintaxis_final = f"{solicitud_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
        if __name__ == "__main__":
            main()

    def CONTENIDO_PALLETS():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Piso Pallets para ver contenido </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Locaciones de Pallets </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de búsqueda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        solicitud_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                        # Crear la parte de la sintaxis para los Nros de WO
                        #solicitud_sintaxis = " | ".join([f'\n{{[Solicitud]:[Nro Solicitud]="{solicitud.strip()}"}}' for solicitud in solicitud_numbers if solicitud.strip()])
                        solicitud_sintaxis = " OR ".join([f'\n Container^LocationCode = "{solicitud.strip()}"' for solicitud in solicitud_numbers if solicitud.strip()])
                        # Construir la sintaxis final
                        sintaxis_final = f"{solicitud_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de búsqueda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
        if __name__ == "__main__":
                main()

    def CODIGOS_DE_6_DIGITOS():
        import pandas as pd
        from io import StringIO

        st.title("📄 Completar códigos a 6 dígitos con ceros")

        st.markdown("""
        Sube un archivo `.txt` que contenga un código por línea.  
        Este script completará cada código a **6 dígitos** rellenando con ceros a la izquierda.
        """)

        uploaded_file = st.file_uploader("📤 Cargar archivo TXT", type=["txt"])

        if uploaded_file is not None:
            # Leer el archivo como texto
            stringio = StringIO(uploaded_file.getvalue().decode("utf-8"))
            lines = stringio.readlines()

            # Eliminar saltos de línea y espacios extra
            codes = [line.strip() for line in lines if line.strip()]

            # Completar con ceros a la izquierda
            completed_codes = [code.zfill(6) for code in codes]

            # Mostrar los primeros 5 códigos
            st.subheader("🔍 Vista previa (primeras 5 filas):")
            preview = completed_codes[:5]
            df_preview = pd.DataFrame(preview, columns=["Código Completado"])
            st.dataframe(df_preview, use_container_width=True)

            # Mostrar todos los resultados (colapsable)
            with st.expander("📋 Ver todos los códigos completados"):
                st.code("\n".join(completed_codes), language='text')

            # Campo para que el usuario ingrese el nombre del archivo de descarga
            filename = st.text_input(
                "✏️ Nombre del archivo de salida (sin extensión):",
                value="codigos_completados"
            )

            # Botón de descarga
            output_text = "\n".join(completed_codes)
            if filename.strip():
                st.download_button(
                    label="⬇️ Descargar archivo procesado",
                    data=output_text,
                    file_name=f"{filename.strip()}.txt",
                    mime="text/plain"
                )
        else:
            st.info("Por favor, sube un archivo `.txt` para procesarlo.")
            
            
        # --------------------- EJECUTAR SEGÚN MENÚ ---------------------
        
    if menu == "🧱 OTPM - MATERIALES":
        OTPM_MATERIALES()
    elif menu == "🗃️ OTEX - EXTRACIONES":
        OTEX_EXTRACIONES()
    elif menu == "📤 OTEV - ENVIO":
        OTEV_ENVIO()
    elif menu == "📥 OTRE - RECOJO":
        OTRE_RECOJO()
    elif menu == "✉️ OTRE - POR_WORKORDERS":
        OTRE_POR_WORKORDERS()
    elif menu == "🧾 OTRE - POR_SOLICITUDES":
        OTRE_POR_SOLICITUDES()
    elif menu == "📝 OTRE - SOLICITUDES":
        OTRE_SOLICITUDES()
    elif menu == "🧠 COD - BASE_DE_DATO":
        COD_BASE_DE_DATO()
    elif menu == "🎨 CONTENIDO_PALLETS":
        CONTENIDO_PALLETS()  
    elif menu == "🔖 CODIGOS DE 6 DIGITOS":
        CODIGOS_DE_6_DIGITOS()   
                        
elif proyectos == "CHECK FILEWEB AND LASERFICHE":
    import streamlit as st
    import pandas as pd
    import os
    from io import BytesIO

    # --------------------- SIDEBAR MENU ---------------------
    st.sidebar.title("Menú de Opciones")
    menu = st.sidebar.radio("Ir a:", [
        "🔎 Generar Sintaxis Solicitudes",
        "🏷️ Reemplazar archivo",
        "🧩 Comparación de Solicitudes FileWeb vs LaserFiche",
        "🔎 Generar Sintaxis LaserFiche del filtrado",
        "🔎 Sintaxis ONEIL del filtrado",
        "🗑️ Eliminar Archivos Generados"

    ])
        
    # --------------------- 1. GENERADOR DE SINTAXIS ---------------------
    def Generar_Sintaxis_Solicitudes():

        st.title("🔎 GENERADOR DE SINTAXIS DE BÚSQUEDA LASER FICHE")

        col1, col2 = st.columns((2))
        with col1:
            # Cargar archivo Excel
            st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCodeFileWeb'</h4>", unsafe_allow_html=True)
            uploaded_file = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")

            # Variable para almacenar la sintaxis generada
            sintaxis_final = ""

            # Generar la sintaxis de búsqueda
            if st.button("Generar Sintaxis"):

                if uploaded_file is not None:
                    try:
                        # Leer el archivo Excel y obtener solo la columna 'SolicitudCode'
                        df = pd.read_excel(uploaded_file)

                        # Verificar si la columna existe
                        if 'SolicitudCodeFileWeb' not in df.columns:
                            st.error("La columna 'SolicitudCodeFileWeb' no se encuentra en el archivo.")
                        else:
                            # Eliminar las demás columnas y eliminar nulos
                            solicitud_codes = df['SolicitudCodeFileWeb'].dropna().astype(str).tolist()

                            # Crear la sintaxis
                            solicitud_sintaxis = " | ".join(
                                [f'\n{{[Orden de Trabajo]:[Nro Solicitud]="{code.strip()}"}}' for code in solicitud_codes if code.strip()]
                                
                            )

                            sintaxis_final = f"{solicitud_sintaxis}"
                    except Exception as e:
                        st.error(f"Error al leer el archivo: {e}")

            # Pie de página
            st.markdown("---")
            st.caption("ℹ️ PARA DESCARGAR DE LASER FICHE DEL CAMPO (Orden de Trabajo) AGREGAR [Nro de Solicitus y Nro de WO] ")

        with col2:
            # Mostrar la sintaxis generada
            if sintaxis_final:
                st.success("Sintaxis de búsqueda generada:")
                st.code(sintaxis_final)

    # --------------------- 2. Reemplazar Archivos ---------------------

    def Reemplazar_Archivo():
        import streamlit as st
        import pandas as pd
        import os
        import io

        #st.set_page_config(page_title="Reemplazo de Archivos Excel", page_icon="📊", layout="wide")
        st.title("Reemplazo de Archivos Excel 🗂️")

        # Función para cargar los archivos Excel
        def cargar_archivo(nombre):
            """Cargar un archivo Excel en un DataFrame y devolver el nombre del archivo."""
            archivo = st.file_uploader(f"Sube el archivo {nombre}", type=["xlsx", "xls"])
            if archivo is not None:
                # Obtener el nombre del archivo
                nombre_archivo = archivo.name
                # Cargar el archivo Excel en un DataFrame de Pandas
                df = pd.read_excel(archivo, engine='openpyxl')
                return df, nombre_archivo
            return None, None

        # Función para reemplazar los datos
        def reemplazar_datos(df1, df2):
            """Reemplazar solo los datos del archivo 1 con los datos del archivo 2, manteniendo el encabezado."""
            # Guardar el encabezado del archivo 1
            encabezado_df1 = df1.columns
            
            # Reemplazar los datos, manteniendo el encabezado del archivo 1
            df2.columns = encabezado_df1  # Aseguramos que los encabezados de df2 coincidan con df1
            df2 = df2[encabezado_df1]     # Solo tomamos las columnas de df2 que están en df1
            return df2

        # Función para guardar el archivo reemplazado en la ruta proporcionada
        def guardar_archivo(df, ruta, nombre_archivo):
            """Guardar el archivo reemplazado en la ruta especificada por el usuario."""
            if not os.path.exists(ruta):
                st.error("🚨 ¡La ruta no existe! Por favor, ingresa una ruta válida.")
                return None
            
            # Guardar el archivo como Excel en la ruta proporcionada
            ruta_completa = os.path.join(ruta, nombre_archivo)
            df.to_excel(ruta_completa, index=False, engine='openpyxl')
            
            return ruta_completa

        # Subir los dos archivos Excel
        col1, col2 = st.columns(2)
        with col1:
            archivo_1_df, nombre_archivo_1 = cargar_archivo("1")
        with col2:
            archivo_2_df, _ = cargar_archivo("2")

        # Solicitar la ruta de guardado
        ruta_guardado = st.text_input("💾 Ingresa la ruta para guardar el archivo reemplazado:", "")

        # Si los archivos han sido subidos
        if archivo_1_df is not None and archivo_2_df is not None:
            # Reemplazar los datos del archivo 1 con los del archivo 2, manteniendo el encabezado de archivo 1
            archivo_reemplazado = reemplazar_datos(archivo_1_df, archivo_2_df)
            
            # Mostrar los primeros 5 registros del archivo reemplazado
            st.subheader("Archivo Reemplazado (Vista Previa) 👀")
            st.write(archivo_reemplazado.head())
            
            if ruta_guardado:
                # Intentar guardar el archivo en la ruta especificada por el usuario
                archivo_guardado = guardar_archivo(archivo_reemplazado, ruta_guardado, nombre_archivo_1)
                
                if archivo_guardado:
                    st.success(f"✅ El archivo fue guardado correctamente en: **{archivo_guardado}**")
            else:
                st.warning("⚠️ Por favor, ingresa una ruta válida para guardar el archivo.")


    # --------------------- 3. Comparación de Solicitudes FileWeb vs LaserFiche ---------------------
    def Comparar_Solicitudes_FW_y_LF():
        import pandas as pd
        import streamlit as st
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
        from collections import defaultdict
        import os

        st.title("📋 Comparación de Solicitudes FileWeb vs LaserFiche")

        def procesar_archivos(file1, file2):
            df1 = pd.read_excel(file1)
            df2 = pd.read_excel(file2)

            # Asegurar que las columnas relevantes sean texto
            columnas_texto = [
                'SolicitudCodeFileWeb', 'Tipo de Solicitud FileWeb', 'Estado de Solicitud FileWeb',
                'WorkOrderCodeFileWe', 'CantidadFileWeb', 'TipoFileFileWeb', 'Centro de Costo FileWeb',
                'Fecha de Registro FileWeb', 'NombreLaserFiche', 'Cliente LaserFiche',
                'Estado de OT LaserFiche', 'Nro Solicitud laserFiche'
            ]
            for df in [df1, df2]:
                for col in columnas_texto:
                    if col in df.columns:
                        df[col] = df[col].astype(str).fillna("")

            merged_df = pd.merge(
                df1[['SolicitudCodeFileWeb', 'Tipo de Solicitud FileWeb', 'Estado de Solicitud FileWeb',
                    'WorkOrderCodeFileWe', 'CantidadFileWeb', 'TipoFileFileWeb', 'Centro de Costo FileWeb',
                    'Fecha de Registro FileWeb']],
                df2[['NombreLaserFiche', 'Cliente LaserFiche', 'Estado de OT LaserFiche', 'Nro Solicitud laserFiche']],
                left_on='SolicitudCodeFileWeb',
                right_on='Nro Solicitud laserFiche',
                how='inner'
            )

            column_order = [
                'SolicitudCodeFileWeb',
                'Nro Solicitud laserFiche',
                'Tipo de Solicitud FileWeb',
                'NombreLaserFiche',
                'Estado de Solicitud FileWeb',
                'Estado de OT LaserFiche',
                'WorkOrderCodeFileWe',
                'Fecha de Registro FileWeb',
                'CantidadFileWeb',
                'TipoFileFileWeb',
                'Centro de Costo FileWeb',
                'Cliente LaserFiche'
            ]

            return merged_df[column_order]

        # Subida de archivos
        col1, col2 = st.columns(2)
        file1 = col1.file_uploader("📂 Sube archivo FileWeb", type=["xlsx"])
        file2 = col2.file_uploader("📂 Sube archivo LaserFiche", type=["xlsx"])

        if file1 and file2:
            try:
                resultado = procesar_archivos(file1, file2)

                # Comparar estados de forma segura
                resultado['Rojo'] = resultado.apply(
                    lambda row: str(row['Estado de Solicitud FileWeb']).strip() != str(row['Estado de OT LaserFiche']).strip(),
                    axis=1
                )

                # Opción de mostrar todo o solo diferentes
                opcion = st.radio("¿Qué archivo deseas ver y descargar?", ("Todo", "Solo Rojo"))

                if opcion == "Solo Rojo":
                    resultado = resultado[resultado['Rojo'] == True]

                # Filtros
                estados_ot = sorted(resultado['Estado de OT LaserFiche'].dropna().unique())
                estado_ot = st.selectbox("📌 Filtrar por Estado de OT LaserFiche:", ["(Todos)"] + estados_ot)

                estados_fileweb = sorted(resultado['Estado de Solicitud FileWeb'].dropna().unique())
                estado_fileweb = st.selectbox("📍 Filtrar por Estado de Solicitud FileWeb:", ["(Todos)"] + estados_fileweb)

                prefijo = st.text_input("🔤 Buscar por prefijo de NombreLaserFiche (ej: OTEV, OTPM):").strip().upper()

                # Aplicar filtros
                if estado_ot != "(Todos)":
                    resultado = resultado[resultado['Estado de OT LaserFiche'] == estado_ot]

                if estado_fileweb != "(Todos)":
                    resultado = resultado[resultado['Estado de Solicitud FileWeb'] == estado_fileweb]

                if prefijo:
                    resultado = resultado[resultado['NombreLaserFiche'].str.upper().str.startswith(prefijo)]

                st.success(f"✅ Se encontraron {len(resultado)} registros filtrados")
                st.dataframe(resultado.drop(columns=['Rojo']), height=600, use_container_width=True)

                # Exportar a Excel con formato
                resultado_exportar = resultado.drop(columns=['Rojo'])
                excel_buffer = BytesIO()
                resultado_exportar.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                wb = load_workbook(excel_buffer)
                ws = wb.active
                ws.auto_filter.ref = ws.dimensions

                # Estilos
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                center_align = Alignment(horizontal="center", vertical="center")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border

                # Mapeo columnas
                header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1)}
                col_estado1 = header_map.get("Estado de Solicitud FileWeb")
                col_estado2 = header_map.get("Estado de OT LaserFiche")
                col_workorder = header_map.get("WorkOrderCodeFileWe")

                verde = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                rojo = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
                intercalados = [PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
                                PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")]

                # Colorear diferencias
                for row in range(2, ws.max_row + 1):
                    val1 = ws.cell(row=row, column=col_estado1).value
                    val2 = ws.cell(row=row, column=col_estado2).value
                    if val1 and val2:
                        if str(val1).strip() != str(val2).strip():
                            ws.cell(row=row, column=col_estado1).fill = rojo
                            ws.cell(row=row, column=col_estado2).fill = rojo
                        else:
                            ws.cell(row=row, column=col_estado1).fill = verde
                            ws.cell(row=row, column=col_estado2).fill = verde

                # Resaltar duplicados
                if col_workorder:
                    duplicados = defaultdict(list)
                    for row in range(2, ws.max_row + 1):
                        val = ws.cell(row=row, column=col_workorder).value
                        if val:
                            duplicados[str(val).strip()].append(row)
                    grupos = [v for v in duplicados.values() if len(v) > 1]
                    for i, filas in enumerate(grupos):
                        fill = intercalados[i % 2]
                        for row in filas:
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = fill

                # Ajustar ancho y aplicar bordes
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = thin_border
                    ws.column_dimensions[col_letter].width = max_length + 2

                final_buffer = BytesIO()
                wb.save(final_buffer)
                final_buffer.seek(0)

                # Guardar localmente si se indica
                ruta_personalizada = st.text_input("📁 Ingrese una ruta de servidor para guardar el archivo (opcional):")
                if ruta_personalizada:
                    try:
                        output_path = os.path.join(ruta_personalizada, "comparacion_de_solicitudes.xlsx")
                        with open(output_path, "wb") as f:
                            f.write(final_buffer.getbuffer())
                        st.success(f"✅ Archivo guardado en: {output_path}")
                    except Exception as e:
                        st.error(f"❌ No se pudo guardar el archivo: {str(e)}")

                st.download_button(
                    label="💾 Descargar archivo con formato",
                    data=final_buffer,
                    file_name="comparacion_de_solicitudes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"❌ Error: {str(e)}")

        st.markdown("---")
        st.caption("🔎 Este archivo contiene filtros, bordes, colores para diferencias y filas duplicadas.")


    def Generar_Sintaxis_del_filtrado():

        st.title("🔎 GENERADOR DE SINTAXIS PARA LASER FICHE DEL ARCHIVO FILTRADO")

        colum1, colum2 = st.columns((2))
        with colum1:
            # Cargar archivo Excel
            st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCodeFileWeb'</h4>", unsafe_allow_html=True)
            uploaded_file = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")

            # Variable para almacenar la sintaxis generada
            sintaxis_final = ""

            # Generar la sintaxis de búsqueda
            if st.button("Generar Sintaxis"):

                if uploaded_file is not None:
                    try:
                        # Leer el archivo Excel y obtener solo la columna 'SolicitudCode'
                        df = pd.read_excel(uploaded_file)

                        # Verificar si la columna existe
                        if 'SolicitudCodeFileWeb' not in df.columns:
                            st.error("La columna 'SolicitudCodeFileWeb' no se encuentra en el archivo.")
                        else:
                            # Eliminar las demás columnas y eliminar nulos
                            solicitud_codes = df['SolicitudCodeFileWeb'].dropna().astype(str).tolist()

                            # Crear la sintaxis
                            solicitud_sintaxis = " | ".join(
                                [f'\n{{[Orden de Trabajo]:[Nro Solicitud]="{code.strip()}"}}' for code in solicitud_codes if code.strip()]
                                
                            )

                            sintaxis_final = f"{solicitud_sintaxis}"
                    except Exception as e:
                        st.error(f"Error al leer el archivo: {e}")

        with colum2:
            # Mostrar la sintaxis generada
            if sintaxis_final:
                st.success("Sintaxis de búsqueda generada:")
                st.code(sintaxis_final)
        
        # Pie de página
        st.markdown("---")

    def Generar_Sintaxis_ONEIL_del_filtrado():
        import streamlit as st
        import pandas as pd

        st.title("🔎 GENERADOR DE SINTAXIS PARA ONEIL DEL ARCHIVO FILTRADO")
        colums1, colums2 = st.columns((2))

        with colums1:
            # Cargar archivo Excel
            st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCodeFileWeb'</h4>", unsafe_allow_html=True)
            uploaded_files = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")

            # Variable para almacenar la sintaxis generada
            sintaxis_final = ""

            if st.button("Generar Sintaxis"):
                if uploaded_files is not None:
                    # Leer el archivo Excel con pandas
                    try:
                        df = pd.read_excel(uploaded_files)

                        # Verificar si la columna existe
                        if "SolicitudCodeFileWeb" in df.columns:
                            # Eliminar valores vacíos y construir la sintaxis
                            solicitud_numbers = df["SolicitudCodeFileWeb"].dropna().astype(str)

                            solicitud_sintaxis = " OR ".join([
                                f'\n Workorder^PONumber = "{solicitud.strip()}"'
                                for solicitud in solicitud_numbers
                            ])
                            sintaxis_final = f"{solicitud_sintaxis}"
                        else:
                            st.error("❌ La columna 'SolicitudCodeFileWeb' no se encontró en el archivo.")
                    except Exception as e:
                        st.error(f"❌ Error al leer el archivo: {e}")

        with colums2:
            if sintaxis_final:
                st.success("Sintaxis de búsqueda generada:")
                st.code(sintaxis_final)

    
   
    # ------************************************************************************------

    def eliminar_archivos_generados():
            import streamlit as st
            import os
            import glob
            import pandas as pd

            #st.set_page_config(page_title="Eliminar Excel con vista previa", layout="centered")
            st.title("🗑️ Eliminar archivos Excel con vista previa desplegable")

            folder_input = st.text_input("Introduce la ruta de la carpeta con archivos Excel:", "C:\\Users\\juan.ramos\\Desktop\\FILTRADOR_ALMACEN\\CHECK_FILEWEB_AND_LASER_FICHE")
            if st.button("🔄 Cargar archivos"):
                st.session_state.folder = folder_input
                st.session_state.selected_files = []

            # Inicializar estados
            if 'folder' not in st.session_state:
                st.session_state.folder = ""
            if 'selected_files' not in st.session_state:
                st.session_state.selected_files = []

            folder = st.session_state.folder

            def toggle_selection(f_name):
                sel = st.session_state.selected_files
                if f_name in sel:
                    sel.remove(f_name)
                else:
                    sel.append(f_name)

            if folder:
                if os.path.isdir(folder):
                    files = glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls"))
                    if files:
                        st.subheader("📂 Selecciona archivos:")
                        for path in files:
                            name = os.path.basename(path)
                            selected = name in st.session_state.selected_files
                            col1, col2 = st.columns([4,1])
                            with col1:
                                st.write(name)
                            with col2:
                                if st.button("✅" if selected else "☐", key=name):
                                    toggle_selection(name)
                                    st.rerun()

                        if st.session_state.selected_files:
                            st.subheader("👁️ Previsualizaciones desplegables")
                            for name in st.session_state.selected_files:
                                file_path = os.path.join(folder, name)
                                with st.expander(f"Ver preview: {name}", expanded=False):
                                    try:
                                        df = pd.read_excel(file_path, nrows=10)
                                        st.dataframe(df)
                                    except Exception as e:
                                        st.error(f"Error leyendo {name}: {e}")

                            if st.checkbox("⚠️ Confirmo que deseo eliminar los archivos seleccionados"):
                                if st.button("🗑️ Eliminar archivos"):
                                    cnt = 0
                                    for name in list(st.session_state.selected_files):
                                        p = os.path.join(folder, name)
                                        if os.path.exists(p):
                                            os.remove(p)
                                            cnt += 1
                                    st.success(f"Se eliminaron {cnt} archivos correctamente.")
                                    st.session_state.selected_files = []
                                    st.session_state.folder = ""
                                    st.rerun()
                        else:
                            st.info("Selecciona archivos pulsando en los botones de la derecha.")
                    else:
                        st.warning("No se encontraron archivos Excel en la carpeta.")
                else:
                    st.error("La ruta no es válida o no corresponde a una carpeta.")
            else:
                st.info("Introduce la ruta y pulsa 'Cargar archivos' para empezar.")


    # --------------------- EJECUTAR SEGÚN MENÚ ---------------------
    if menu == "🔎 Generar Sintaxis Solicitudes":
        Generar_Sintaxis_Solicitudes()
    elif menu == "🧩 Comparación de Solicitudes FileWeb vs LaserFiche":
        Comparar_Solicitudes_FW_y_LF()
    elif menu == "🏷️ Reemplazar archivo":
        Reemplazar_Archivo()
    elif menu == "🔎 Generar Sintaxis LaserFiche del filtrado":
        Generar_Sintaxis_del_filtrado()
    elif menu == "🔎 Sintaxis ONEIL del filtrado":
        Generar_Sintaxis_ONEIL_del_filtrado()
    elif menu == "🗑️ Eliminar Archivos Generados":
        eliminar_archivos_generados()

elif proyectos == "SCRIPTS":

    import streamlit as st
    from PyPDF2 import PdfWriter, PdfReader
    import io
    import os

    st.sidebar.header("SELECCIONES UNA OPCION: ")
    scripts = st.sidebar.selectbox(
        "Opciones",
        options=["OPCIONES", "UNIR PDF", "MOSTRAR CODIGOS", "COMPRIMIR ARCHIVOS", "UNIR 2 PAGINAS PDF", "SINTAXIS LASERFICHE", "CREAR RUTAS", "DUPLICAR FILAS", "FILEWEB AND LASERFICHE"]
    )
    if scripts == "OPCIONES":
        st.title(" :card_index_dividers: SCRIPTS CREADOS POR JUAN CARLOS RAMOS")
        st.markdown('<style>div.block-container{padding-top:2rem;}</style>',unsafe_allow_html=True)

        # Mostrar el total de ingresos centrado en una caja
        st.markdown(
            f"""
            <div style='
                border: 2px solid black; 
                padding: 2px; 
                border-radius: 2px; 
                text-align: center; 
                background-color: #3b3938;'>
                <h3 style='color: white;'>BIENVENIDO A SCRIPTS CREADOS PARA FACILITAR TAREAS</h3>
                
            </div>
            """, 
            unsafe_allow_html=True
        )

        # Mostrar la imagen de portada
        st.image("C:/Users/juan.ramos/Desktop/FILTRADOR_ALMACEN/Imagenes/pyscript.jpg", use_column_width=True, caption="Derechos-recervados-@JuanCarlos-Ramos-2024 ")


    # ------************************************************************************------
    elif scripts == "UNIR PDF": 

        def main():
            st.title("Unir Archivos PDF")

            # Cargar archivos PDF
            uploaded_files = st.file_uploader("Selecciona los archivos PDF", type="pdf", accept_multiple_files=True)

            # Ingresar una ruta donde se guardara el archivo unido
            folder_path = st.text_input("Ingresauna ruta para guardar el archivo: ")

            if uploaded_files and folder_path:
                # Crear un objeto PdfWriter
                pdf_writer = PdfWriter()

                # Leer cada archivo PDF y agregarlo al PdfWriter
                for pdf_file in uploaded_files:
                    pdf_reader = PdfReader(pdf_file)
                    for page in range(len(pdf_reader.pages)):
                        pdf_writer.add_page(pdf_reader.pages[page])

                # Crear un buffer en memoria para guardar el PDF combinado
                output_pdf = io.BytesIO()
                pdf_writer.write(output_pdf)
                output_pdf.seek(0)

                # Mostrar un mensaje de texto
                st.success("Los archivos PDF se han unido correctamente.")

                # Guardar el PDF combinado e la ruta especifica
                output_file_path = os.path.join(folder_path, "pdf_combinado.pdf")
                with open(output_file_path, "wb") as f:
                    f.write(output_pdf.getbuffer())


                # Descargar el PDF combinado
                st.download_button(
                    label="Descargar PDF Combinado",
                    data=output_pdf,
                    file_name="pdf_combinado.pdf",
                    mime="application/pdf"
                )

        if __name__ == "__main__":
            main()
    # ------************************************************************************------

    # ------ *******************OTHER SCRIPT****************************************------

    # ------************************************************************************------    
    elif scripts == "MOSTRAR CODIGOS":
        import streamlit as st
        import pandas as pd
        from io import BytesIO

        # Título de la aplicación
        st.title("FILTRAR CODIGOS POLYSISTEMAS")

        # Subida de archivo
        uploaded_file = st.file_uploader("Sube un archivo Excel", type=["xlsx", "xls"])

        if uploaded_file is not None:
            # Leer el archivo Excel
            df = pd.read_excel(uploaded_file, dtype=str)  # Leer todo como texto para preservar ceros
            
            # Mostrar el DataFrame original
            st.write("DataFrame Original:")
            st.dataframe(df)

            # Verificar si la columna "Items Oneil" existe
            if "Items Oneil" in df.columns:
                # Seleccionar solo la columna "Items Oneil"
                df = df[["Items Oneil"]]
                
                # Crear una lista para almacenar los códigos preservando ceros
                all_codes = []

                # Iterar sobre cada celda en la columna "Items Oneil"
                for value in df["Items Oneil"]:
                    if pd.notna(value):  # Verificar si el valor no es nulo
                        if isinstance(value, str):  # Si es cadena
                            # Separar por comas, eliminar espacios EXTERNOS pero mantener los internos
                            codes = [code.strip() for code in str(value).split(',')]
                            all_codes.extend([code for code in codes if code])  # Filtrar strings vacíos
                
                # Eliminar duplicados manteniendo el orden y los ceros
                seen = set()
                unique_codes = [x for x in all_codes if x and not (x in seen or seen.add(x))]

                # Crear un DataFrame con los códigos únicos
                codes_df = pd.DataFrame(unique_codes, columns=["Códigos"])
                
                # Mostrar el DataFrame con los códigos únicos
                st.write("Códigos Únicos (con ceros a la izquierda):")
                st.dataframe(codes_df)

                # Opción para descargar el resultado en Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    codes_df.to_excel(writer, index=False, sheet_name='Códigos')
                
                output.seek(0)
                st.download_button(
                    label="Descargar Códigos Únicos en Excel",
                    data=output,
                    file_name="codigos_unicos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("La columna 'Items Oneil' no se encuentra en el archivo.")

    # ------************************************************************************------

    # ------ *******************OTHER SCRIPT****************************************------
    
    # ------************************************************************************------
    elif scripts == "COMPRIMIR ARCHIVOS":
        # pip install streamlit pyzipper
        import streamlit as st
        import zipfile
        import pyzipper
        import os
        from io import BytesIO

        # Interfaz de usuario de Streamlit
        st.title("COMPRESOR DE ARCHIVOS")

         # Pedimos al usuario que elija una opcion
        Comprimir = st.selectbox(
            "Opciones",
            options=["COMPRIMIR", "SIN CONTRASEÑA", "CON CONTRASEÑA"]
        )

        if Comprimir == "SIN CONTRASEÑA":

            # Función para comprimir los archivos en un archivo ZIP
            def compress_files(files):
                zip_buffer = BytesIO()
                
                # Crear el archivo ZIP en memoria
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for file in files:
                        file_name = file.name
                        # Agregar el archivo al ZIP, manteniendo su nombre original
                        zip_file.writestr(file_name, file.getvalue())
                
                # Mover el puntero al principio para que pueda ser descargado
                zip_buffer.seek(0)
                return zip_buffer
            
            # Interfaz de usuario de Streamlit
            st.title("Compresor de Archivos sin contraseña")

            # Explicación
            st.write("Sube tus archivos para comprimirlos en un archivo ZIP.")

            # Subir múltiples archivos
            uploaded_files = st.file_uploader("Selecciona los archivos", accept_multiple_files=True)

            if uploaded_files:
                # Mostrar el nombre de los archivos subidos
                st.write("Archivos seleccionados:")
                for uploaded_file in uploaded_files:
                    st.write(f"- {uploaded_file.name}")
                    
                    # Comprimir los archivos cuando el usuario hace clic en el botón
                    if st.button("Comprimir Archivos"):
                        # Llamar a la función para comprimir los archivos
                        zip_buffer = compress_files(uploaded_files)
                        
                        # Crear un archivo comprimido para la descarga
                        st.download_button(
                            label="Descargar archivo ZIP",
                            data=zip_buffer,
                            file_name="archivos_comprimidos.zip",
                            mime="application/zip"
                        )

        
        if Comprimir == "CON CONTRASEÑA":
            # pip install --upgrade pyzipper
            pass
    # ------************************************************************************------

    # ------ *******************OTHER SCRIPT****************************************------
    
    # ------************************************************************************------      
    elif scripts == "UNIR 2 PAGINAS PDF":
        import streamlit as st
        import PyPDF2
        from io import BytesIO

        def unir_dos_paginas(pdf_file, pagina1, pagina2):
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            pdf_writer = PyPDF2.PdfWriter()

            # Asegurarse de que las páginas seleccionadas estén dentro del rango
            if pagina1 < len(pdf_reader.pages) and pagina2 < len(pdf_reader.pages):
                pdf_writer.add_page(pdf_reader.pages[pagina1])
                pdf_writer.add_page(pdf_reader.pages[pagina2])
            else:
                st.error("Las páginas seleccionadas están fuera del rango del PDF.")
                return None

            # Crear un objeto BytesIO para guardar el PDF combinado en memoria
            pdf_salida = BytesIO()
            pdf_writer.write(pdf_salida)
            pdf_salida.seek(0)  # Volver al inicio del objeto BytesIO
            return pdf_salida

        # Título de la aplicación
        st.title("Unir Dos Páginas de un PDF")

        # Cargar archivo PDF
        pdf_file = st.file_uploader("Selecciona un archivo PDF", type=["pdf"])

        if pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_paginas = len(pdf_reader.pages)

            # Mostrar el número total de páginas
            st.write(f"El PDF tiene {num_paginas} páginas.")

            # Seleccionar las páginas a unir
            pagina1 = st.number_input("Selecciona la primera página (0 a {}):".format(num_paginas - 1), min_value=0, max_value=num_paginas - 1, value=0)
            pagina2 = st.number_input("Selecciona la segunda página (0 a {}):".format(num_paginas - 1), min_value=0, max_value=num_paginas - 1, value=1)

            if st.button("Unir Páginas"):
                # Unir las dos páginas seleccionadas
                pdf_combinado = unir_dos_paginas(pdf_file, pagina1, pagina2)

                if pdf_combinado:
                    # Descargar el PDF combinado
                    st.download_button(
                        label="Descargar PDF Combinado",
                        data=pdf_combinado,
                        file_name="pdf_combinado.pdf",
                        mime="application/pdf"
                    )
    # ------************************************************************************------

    # ------ *******************OTHER SCRIPT****************************************------
    
    # ------************************************************************************------
    elif scripts == "DUPLICAR FILAS":
        import streamlit as st
        import pandas as pd
        from io import BytesIO

        # Título de la aplicación
        st.title("Duplicar Filas de un Archivo Excel")

        # Cargar archivo Excel
        uploaded_file = st.file_uploader("Elige un archivo Excel", type=["xlsx"])

        if uploaded_file is not None:
            # Leer el archivo Excel
            df = pd.read_excel(uploaded_file)

            # Mostrar el DataFrame original
            st.write("DataFrame Original:")
            st.dataframe(df)

            # Duplicar filas
            if st.button("Duplicar Filas"):
                # Crear un nuevo DataFrame duplicando cada fila
                duplicated_rows = pd.DataFrame(columns=df.columns)
                
                for index, row in df.iterrows():
                    duplicated_rows = pd.concat([duplicated_rows, pd.DataFrame([row])], ignore_index=True)
                    duplicated_rows = pd.concat([duplicated_rows, pd.DataFrame([row])], ignore_index=True)

                st.write("DataFrame Duplicado:")
                st.dataframe(duplicated_rows)

                # Opción para descargar el archivo duplicado como Excel
                def to_excel(df):
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Duplicado')
                    return output.getvalue()

                excel_data = to_excel(duplicated_rows)
                st.download_button("Descargar archivo duplicado como Excel", excel_data, "duplicado.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ------************************************************************************------

    # ------ *******************OTHER SCRIPT****************************************------
    
    # ------************************************************************************------
    elif scripts == "SINTAXIS LASERFICHE":
        import streamlit as st
        from datetime import datetime


        # Configurar la página en modo ancho
        #st.set_page_config(layout="wide")

        # Título de la aplicación
        st.markdown("<h1 style='text-align: center;'>SINTAXIS PARA LASER FICHE </h1>", unsafe_allow_html=True)


        # Crear una lista de opciones para el menu 
        opciones = ["SELECCIONA UNA OPCION", "OTPM - MATERIALES", "OTEX - EXTRACIONES", "OTEV - ENVIO", "OTRE - RECOJO", "OTRE - POR_WORKORDERS", "OTRE - POR_SOLICITUDES", "OTRE - SOLICITUDES", "COD - BASE_DE_DATOS", "CONTENIDO_PALLETS"]

        # Crear el menu desplegable con st.selectbox()
        seleccion = st.selectbox("Selecciona una opcion del menu: ", opciones)

        if seleccion== "OTPM - MATERIALES":

            # Crear una lista de opciones para el menu 
            opciones2 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

            if seleccion2 == "PENDIENTE":
                import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\03. Finalizada\\01. Pendiente"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion2 == "EN PROCESO":
                import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\02. En Proceso"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion2 == "FINALIZADA":
                import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\03. Finalizada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

            elif seleccion2 == "ANULADA":
                import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparación de Materiales\\04. Anulada"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

        if seleccion== "OTEX - EXTRACIONES":

            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\01. Pendiente"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()


            elif seleccion3 == "EN PROCESO":
                import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\02. En proceso"}'

                    # Sección para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                if __name__ == "__main__":
                    main()

                elif seleccion3 == "FINALIZADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\03. Finalizada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion3 == "ANULADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracción\\04. Anulada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()
                                
        if seleccion== "OTEV - ENVIO":

            # Crear una lista de opciones para el menu 
            opciones2 = ["SELECCIONA UNA OPCION", "INTERNO", "EXTERNO"]

            # Crear el menu desplegable con st.selectbox()
            seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

            if seleccion2 == "INTERNO":
                # Crear una lista de opciones para el menu 
                opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

                # Crear el menu desplegable con st.selectbox()
                seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

                if seleccion3 == "PENDIENTE":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\01. Pendiente"}'

                        # Sección para seleccionar una fecha usando un calendariO
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()


                elif seleccion2 == "EN PROCESO":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\02. En proceso"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion2 == "FINALIZADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\03. Finalizada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion2 == "ANULADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Interno\\04. Anulada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

            if seleccion2 == "EXTERNO":
                # Crear una lista de opciones para el menu 
                opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

                # Crear el menu desplegable con st.selectbox()
                seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

                if seleccion3 == "PENDIENTE":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\01. Pendiente"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()


                elif seleccion3 == "EN PROCESO":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\02. En proceso"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion3 == "FINALIZADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\03. Finalizada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion3 == "ANULADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Envío\\01. Solicitante Externo\\04. Anulada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

        if seleccion== "OTRE - RECOJO":

            # Crear una lista de opciones para el menu 
            opciones2 = ["SELECCIONA UNA OPCION", "INTERNO", "EXTERNO"]

            # Crear el menu desplegable con st.selectbox()
            seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

            if seleccion2 == "INTERNO":
                # Crear una lista de opciones para el menu 
                opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

                # Crear el menu desplegable con st.selectbox()
                seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

                if seleccion3 == "PENDIENTE":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\01. Pendiente"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.header("Seleccione una fecha")
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()


                elif seleccion3 == "EN PROCESO":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\02. En proceso"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion3 == "FINALIZADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\03. Finalizada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion3 == "ANULADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\04. Anulada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

            if seleccion2 == "EXTERNO":
                # Crear una lista de opciones para el menu 
                opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

                # Crear el menu desplegable con st.selectbox()
                seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

                if seleccion3 == "PENDIENTE":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\01. Pendiente"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()


                elif seleccion3 == "EN PROCESO":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\02. En proceso"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion3 == "FINALIZADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\03. Finalizada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

                elif seleccion3 == "ANULADA":
                    import streamlit as st
                    from datetime import datetime

                    def main():
                        st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                        # Sintaxis predeterminada
                        sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\04. Anulada"}'

                        # Sección para seleccionar una fecha usando un calendario
                        st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                        fecha_input = st.date_input("Fecha de creación:", value=datetime.today())

                        # Variable para almacenar la sintaxis generada
                        sintaxis_final = ""

                        # Generar la sintaxis de búsqueda
                        if st.button("Generar Sintaxis"):
                            # Formatear la fecha seleccionada
                            fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                            sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                            # Mostrar la sintaxis generada
                            st.success("Sintaxis de búsqueda generada:")
                            st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                    if __name__ == "__main__":
                        main()

        if seleccion== "OTRE - POR_WORKORDERS":
            import streamlit as st
            from datetime import datetime

            def main():
                col1, col2 = st.columns((2))
                with col1:

                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Cargar archivo de texto
                    st.markdown("<h4>Selecciona el archivo de texto con Nros de WOs </h4>", unsafe_allow_html=True)
                    uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        
                        # Leer el archivo y extraer los Nros de WO
                        if uploaded_file is not None:
                            content = uploaded_file.read().decode("utf-8")
                            wo_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                            # Crear la parte de la sintaxis para los Nros de WO
                            wo_sintaxis = " | ".join([f'\n{{[Orden de Trabajo]:[Nro de WO]="{wo.strip()}"}}' for wo in wo_numbers if wo.strip()])

                            # Construir la sintaxis final
                            sintaxis_final = f"{wo_sintaxis}"

                with col2:

                    # Mostrar la sintaxis generada
                    if sintaxis_final:
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
            if __name__ == "__main__":
                main()

        if seleccion== "OTRE - POR_SOLICITUDES":
            import streamlit as st
            from datetime import datetime

            def main():
                col1, col2 = st.columns((2))
                with col1:

                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Cargar archivo de texto
                    st.markdown("<h4>Selecciona el archivo de texto con Nros de Solicitudes </h4>", unsafe_allow_html=True)
                    uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        
                        # Leer el archivo y extraer los Nros de WO
                        if uploaded_file is not None:
                            content = uploaded_file.read().decode("utf-8")
                            wo_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                            # Crear la parte de la sintaxis para los Nros de WO
                            wo_sintaxis = " | ".join([f'\n{{[Orden de Trabajo]:[Nro Solicitud]="{wo.strip()}"}}' for wo in wo_numbers if wo.strip()])

                            # Construir la sintaxis final
                            sintaxis_final = f"{wo_sintaxis}"

                with col2:

                    # Mostrar la sintaxis generada
                    if sintaxis_final:
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
            if __name__ == "__main__":
                main()

        if seleccion =="OTRE - SOLICITUDES":
            import streamlit as st
            from datetime import datetime

            def main():
                col1, col2 = st.columns((2))
                with col1:

                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Cargar archivo de texto
                    st.markdown("<h4>Selecciona el archivo de texto con Nros de SOLICITUDES </h4>", unsafe_allow_html=True)
                    uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        
                        # Leer el archivo y extraer los Nros de WO
                        if uploaded_file is not None:
                            content = uploaded_file.read().decode("utf-8")
                            solicitud_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                            # Crear la parte de la sintaxis para los Nros de WO
                            solicitud_sintaxis = " | ".join([f'\n{{[Solicitud]:[Nro Solicitud]="{solicitud.strip()}"}}' for solicitud in solicitud_numbers if solicitud.strip()])

                            # Construir la sintaxis final
                            sintaxis_final = f"{solicitud_sintaxis}"

                with col2:

                    # Mostrar la sintaxis generada
                    if sintaxis_final:
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
            if __name__ == "__main__":
                main()

        if seleccion == "COD - BASE_DE_DATOS":
            import streamlit as st
            from datetime import datetime

            def main():
                col1, col2 = st.columns((2))
                with col1:

                    st.markdown("<h3>Generador de Codigos para descargar BD </h3>", unsafe_allow_html=True)

                    # Cargar archivo de texto
                    st.markdown("<h4>Selecciona el archivo de texto con Codigos Poly </h4>", unsafe_allow_html=True)
                    uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        
                        # Leer el archivo y extraer los Nros de WO
                        if uploaded_file is not None:
                            content = uploaded_file.read().decode("utf-8")
                            solicitud_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                            # Crear la parte de la sintaxis para los Nros de WO
                            #solicitud_sintaxis = " | ".join([f'\n{{[Solicitud]:[Nro Solicitud]="{solicitud.strip()}"}}' for solicitud in solicitud_numbers if solicitud.strip()])
                            solicitud_sintaxis = " OR ".join([f'\n Filefolder^ContainerCode = "{solicitud.strip()}"' for solicitud in solicitud_numbers if solicitud.strip()])
                            # Construir la sintaxis final
                            sintaxis_final = f"{solicitud_sintaxis}"

                with col2:

                    # Mostrar la sintaxis generada
                    if sintaxis_final:
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
            if __name__ == "__main__":
                main()

        if seleccion == "CONTENIDO_PALLETS":
            import streamlit as st
            from datetime import datetime

            def main():
                col1, col2 = st.columns((2))
                with col1:

                    st.markdown("<h3>Generador de Piso Pallets para ver contenido </h3>", unsafe_allow_html=True)

                    # Cargar archivo de texto
                    st.markdown("<h4>Selecciona el archivo de texto con Locaciones de Pallets </h4>", unsafe_allow_html=True)
                    uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de búsqueda
                    if st.button("Generar Sintaxis"):
                        
                        # Leer el archivo y extraer los Nros de WO
                        if uploaded_file is not None:
                            content = uploaded_file.read().decode("utf-8")
                            solicitud_numbers = content.splitlines()  # Suponiendo que cada WO está en una línea

                            # Crear la parte de la sintaxis para los Nros de WO
                            #solicitud_sintaxis = " | ".join([f'\n{{[Solicitud]:[Nro Solicitud]="{solicitud.strip()}"}}' for solicitud in solicitud_numbers if solicitud.strip()])
                            solicitud_sintaxis = " OR ".join([f'\n Container^LocationCode = "{solicitud.strip()}"' for solicitud in solicitud_numbers if solicitud.strip()])
                            # Construir la sintaxis final
                            sintaxis_final = f"{solicitud_sintaxis}"

                with col2:

                    # Mostrar la sintaxis generada
                    if sintaxis_final:
                        st.success("Sintaxis de búsqueda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de código

                        
            if __name__ == "__main__":
                main()

    # ------************************************************************************------

    # ------ *******************OTHER SCRIPT****************************************------
    
    # ------************************************************************************------
    elif scripts == "CREAR RUTAS":
        import streamlit as st
        import pandas as pd

        # Título de la aplicación
        st.title("PREPARA RUTA ALMACEN LA PAZ")

        col1, col2 = st.columns(2)
        with col1:
            # Cargar el primer archivo Excel
            file1 = st.file_uploader("Cargar el primer archivo Excel", type=["xlsx"])
        with col2:
            # Cargar el segundo archivo Excel
            file2 = st.file_uploader("Cargar el segundo archivo Excel", type=["xlsx"])

        if file1 and file2:
            # Leer los archivos Excel
            df1 = pd.read_excel(file1)
            df1_mod = df1.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8','ELIMINAR_9'], axis=1)
            df2 = pd.read_excel(file2)
            with col1:
                # Mostrar los DataFrames cargados
                st.write("Contenido del primer archivo:")
                st.dataframe(df1_mod)
            with col2:
                st.write("Contenido del segundo archivo:")
                st.dataframe(df2)

            # Definir las columnas que contienen los códigos de solicitud y el solicitante
            request_code_column_df1 = 'SolicitudCode'  # Columna en el primer archivo
            request_code_column_df2 = 'Nro Solicitud'   # Columna en el segundo archivo
            requester_column_df2 = 'Solicitante'         # Columna en el segundo archivo
            user_column_df2 = 'Usuario'
            items_column_df2 = 'Items Oneil'
            centro_costos_column = 'Centro de Costo'

            # Verificar si las columnas existen
            if (request_code_column_df1 in df1_mod.columns and request_code_column_df2 in df2.columns and requester_column_df2 in df2.columns and
            user_column_df2 in df2.columns and
            items_column_df2 in df2.columns):
                # Realizar la comparación y agregar la columna de solicitante
                merged_df = pd.merge(df1_mod, df2[[request_code_column_df2, requester_column_df2, user_column_df2, items_column_df2]], 
                                    left_on=request_code_column_df1,
                                    right_on=request_code_column_df2, 
                                    how='left')
                

                if centro_costos_column in merged_df.columns:
                    merged_df = merged_df.sort_values(by=centro_costos_column)
                else:
                    st.warning(f'La Columna "{centro_costos_column}" no existe en el dataFrame combianado')
                    
                #merged_df.columns = ['SolicitudCode', 'Nro Solicitud', 'Tipo de Solicitud', 'Cliente', 'WorkOrderCode', 'Cantidad', 'TipoFile','Centro de Costos', 'Solicitante', 'Usuario', 'Items Oneil']

                #merged_df = merged_df.sort_values(by='Centro de costos')
                # Mostrar el DataFrame resultante
                st.write("DataFrame combinado con la columna de Solicitante:")
                st.dataframe(merged_df)

                # Opción para descargar el DataFrame combinado como archivo Excel
                excel_file = "archivo_combinado.xlsx"
                merged_df.to_excel(excel_file, index=False)

                # Crear un botón de descarga
                with open(excel_file, "rb") as f:
                    st.download_button(
                        label="Descargar archivo combinado",
                        data=f,
                        file_name=excel_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            else:
                st.error("Una o más columnas especificadas no existen en los DataFrames.")

    # ------************************************************************************------

    # ------ *******************OTHER SCRIPT****************************************------
    
    # ------************************************************************************------
    
    elif scripts == "FILEWEB AND LASERFICHE":
        import pandas as pd
        import streamlit as st
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        # Configuración de la página
        #st.set_page_config(page_title="Comparador de Solicitudes", layout="wide")
        st.title("📋 Comparación de Solicitudes FileWeb vs LaserFiche")

        # Función para procesar y comparar los archivos
        def procesar_archivos(file1, file2):
            df1 = pd.read_excel(file1)
            df2 = pd.read_excel(file2)
            
            merged_df = pd.merge(
                df1[['SolicitudCodeFileWeb', 'Tipo de Solicitud FileWeb', 'Estado de Solicitud FileWeb',
                    'WorkOrderCodeFileWe', 'CantidadFileWeb', 'TipoFileFileWeb', 'Centro de Costo FileWeb',
                    'Fecha de Registro FileWeb']],
                df2[['Cliente LaserFiche', 'Estado de OT LaserFiche', 'Nro Solicitud laserFiche']],
                left_on='SolicitudCodeFileWeb',
                right_on='Nro Solicitud laserFiche',
                how='inner'
            )
            
            column_order = [
                'SolicitudCodeFileWeb',
                'Nro Solicitud laserFiche',
                'Tipo de Solicitud FileWeb',
                'Estado de Solicitud FileWeb',
                'Estado de OT LaserFiche',
                'WorkOrderCodeFileWe',
                'Fecha de Registro FileWeb',
                'CantidadFileWeb',
                'TipoFileFileWeb',
                'Centro de Costo FileWeb',
                'Cliente LaserFiche'
            ]
            
            return merged_df[column_order]

        # Interfaz de usuario
        col1, col2 = st.columns(2)

        with col1:
            file1 = st.file_uploader("Sube archivo FileWeb", type=["xlsx"], key="file1")

        with col2:
            file2 = st.file_uploader("Sube archivo LaserFiche", type=["xlsx"], key="file2")

        if file1 and file2:
            try:
                resultado = procesar_archivos(file1, file2)
                st.success(f"✅ Se encontraron {len(resultado)} coincidencias")

                # Mostrar tabla estilizada en Streamlit
                def resaltar_estados(row):
                    if str(row['Estado de Solicitud FileWeb']).strip() == str(row['Estado de OT LaserFiche']).strip():
                        return [''] * 2 + ['background-color: lightgreen'] * 2 + [''] * (len(row) - 4)
                    else:
                        return [''] * 2 + ['background-color: lightcoral'] * 2 + [''] * (len(row) - 4)

                styled_df = resultado.style\
                    .apply(resaltar_estados, axis=1)\
                    .set_properties(**{'background-color': '#f9f9f9', 'color': '#333'})\
                    .set_table_styles([{
                        'selector': 'th',
                        'props': [('background-color', '#4a7dff'), ('color', 'white')]
                    }])

                st.dataframe(styled_df, height=600, use_container_width=True)

                # Guardar DataFrame en un buffer Excel
                excel_buffer = BytesIO()
                resultado.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                # Cargar con openpyxl para aplicar estilos
                wb = load_workbook(excel_buffer)
                ws = wb.active

                # Congelar encabezado y agregar filtro
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                # Detectar columnas por nombre
                header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1)}
                col_estado_fileweb = header_map.get("Estado de Solicitud FileWeb")
                col_estado_laserfiche = header_map.get("Estado de OT LaserFiche")

                if col_estado_fileweb and col_estado_laserfiche:
                    verde = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    rojo = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")

                    for row in range(2, ws.max_row + 1):
                        val1 = ws.cell(row=row, column=col_estado_fileweb).value
                        val2 = ws.cell(row=row, column=col_estado_laserfiche).value
                        if val1 is not None and val2 is not None:
                            if str(val1).strip() == str(val2).strip():
                                ws.cell(row=row, column=col_estado_fileweb).fill = verde
                                ws.cell(row=row, column=col_estado_laserfiche).fill = verde
                            else:
                                ws.cell(row=row, column=col_estado_fileweb).fill = rojo
                                ws.cell(row=row, column=col_estado_laserfiche).fill = rojo

                # Guardar archivo final
                final_buffer = BytesIO()
                wb.save(final_buffer)
                final_buffer.seek(0)

                # Botón de descarga
                st.download_button(
                    label="💾 Descargar Archivo Filtrado",
                    data=final_buffer,
                    file_name="comparacion_solicitudes_coloreado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except KeyError as e:
                st.error(f"Error: No se encontró la columna {str(e)} en uno de los archivos")
            except Exception as e:
                st.error(f"Error al procesar los archivos: {str(e)}")

        # Pie de página
        st.markdown("---")
        st.caption("ℹ️ Este comparador muestra las solicitudes que existen tanto en FileWeb como en LaserFiche, resaltando los estados coincidentes en verde y los diferentes en rojo.")

    # ------************************************************************************------

    
